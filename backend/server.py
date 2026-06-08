"""FastAPI server for the Night Shift Reports automation.

Exposes endpoints that the Vite frontend consumes via its /api proxy.
"""
from __future__ import annotations

import shutil
import uuid
import zipfile
from io import BytesIO
from pathlib import Path
from typing import Any

from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import FileResponse

from openpyxl import load_workbook

from nightshift import constants as K
from nightshift.part1_get_results import compile_all_orders, extract_jetro_and_po
from nightshift.part1_shortages import calculate_shortages
from nightshift.part1_setup import setup_sheets_and_dropdowns
from nightshift.part2_apply_routing import apply_routing
from nightshift.part3_final_pick_sheets import build_final_pick_sheets
from nightshift.part4_jetro_workflow import build_jetro_branch
from nightshift.sheet_utils import find_sheet, header_map, iter_data_rows, normalize_header_typos

# ---------------------------------------------------------------------------
# App + job storage
# ---------------------------------------------------------------------------

JOBS_DIR = Path(__file__).parent / "_jobs"
JOBS_DIR.mkdir(exist_ok=True)

app = FastAPI(title="Night Shift Reports API")


def _job_dir(job_id: str) -> Path:
    p = JOBS_DIR / job_id
    p.mkdir(parents=True, exist_ok=True)
    return p


# ---------------------------------------------------------------------------
# Helpers — preview extraction
# ---------------------------------------------------------------------------

def _extract_preview(wb: Any) -> dict[str, Any]:
    """Build the preview payload the frontend expects after Part 1."""

    # All Orders
    try:
        ao = find_sheet(wb, K.SHEET_ALL_ORDERS)
        known_ao = ["Product Name", "Code", "Bin", "Vendor", "Sheet", "BIN(Internal)", "BIN (Internal)", "internal bin", "BIN"]
        ao_hmap = header_map(ao, known_headers=known_ao)
        ao_row_count = max(ao.max_row - 1, 0)
        ao_rows = []
        for row_idx, values in iter_data_rows(ao, known_headers=known_ao):
            # Try to get internal bin for preview
            ib_idx = ao_hmap.get("BIN(Internal)") or ao_hmap.get("BIN (Internal)") or ao_hmap.get("internal bin") or ao_hmap.get("BIN")
            ib_val = str(values[ib_idx-1]) if ib_idx and ib_idx <= len(values) else ""
            
            ao_name_idx = ao_hmap.get("Name")
            ao_rows.append({
                "id": f"{K.SHEET_ALL_ORDERS}:{row_idx}",
                "values": values,
                "productName": values[ao_hmap["Product Name"] - 1] if "Product Name" in ao_hmap else "",
                "code": values[ao_hmap["Code"] - 1] if "Code" in ao_hmap else "",
                "bin": ib_val or (values[ao_hmap["Bin"] - 1] if "Bin" in ao_hmap else ""),
                "vendor": values[ao_hmap["Vendor"] - 1] if "Vendor" in ao_hmap else "",
                "customer": values[ao_name_idx - 1] if ao_name_idx and ao_name_idx <= len(values) else "",
                "sheet": values[0], # Col A
                "vendorRoute": values[1], # Col B
            })
    except KeyError:
        ao_row_count = 0
        ao_rows = []

    # Jetro Source — display the numeric Bin (column F) per PRD §6.4, not the
    # warehouse internal bin. Jetro rows are defined by a numeric Bin so this
    # is the more useful identifier in the routing UI.
    try:
        js = find_sheet(wb, K.SHEET_JETRO_SOURCE)
        known_js = ["Product Name", "Code", "Bin", "Vendor", "Sheet"]
        js_hmap = header_map(js, known_headers=known_js)
        js_row_count = max(js.max_row - 1, 0)
        js_rows = []
        for row_idx, values in iter_data_rows(js, known_headers=known_js):
            bin_val = str(values[js_hmap["Bin"] - 1]) if "Bin" in js_hmap else ""

            js_name_idx = js_hmap.get("Name")
            js_rows.append({
                "id": f"{K.SHEET_JETRO_SOURCE}:{row_idx}",
                "values": values,
                "productName": values[js_hmap["Product Name"] - 1] if "Product Name" in js_hmap else "",
                "code": values[js_hmap["Code"] - 1] if "Code" in js_hmap else "",
                "bin": bin_val,
                "vendor": values[js_hmap["Vendor"] - 1] if "Vendor" in js_hmap else "",
                "customer": values[js_name_idx - 1] if js_name_idx and js_name_idx <= len(values) else "",
                "sheet": values[0],
                "vendorRoute": values[1],
            })
    except KeyError:
        js_row_count = 0
        js_rows = []

    # PO
    try:
        po = find_sheet(wb, K.SHEET_PO)
        known_po = ["Product Name", "Code", "Bin", "Vendor", "Sheet", "BIN(Internal)", "BIN (Internal)", "internal bin", "BIN"]
        po_hmap = header_map(po, known_headers=known_po)
        po_row_count = max(po.max_row - 1, 0)
        po_rows = []
        for row_idx, values in iter_data_rows(po, known_headers=known_po):
            # Try to get internal bin for preview
            ib_idx = po_hmap.get("BIN(Internal)") or po_hmap.get("BIN (Internal)") or po_hmap.get("internal bin") or po_hmap.get("BIN")
            ib_val = str(values[ib_idx-1]) if ib_idx and ib_idx <= len(values) else ""
            
            po_name_idx = po_hmap.get("Name")
            po_rows.append({
                "id": f"{K.SHEET_PO}:{row_idx}",
                "values": values,
                "productName": values[po_hmap["Product Name"] - 1] if "Product Name" in po_hmap else "",
                "code": values[po_hmap["Code"] - 1] if "Code" in po_hmap else "",
                "bin": ib_val or (values[po_hmap["Bin"] - 1] if "Bin" in po_hmap else ""),
                "vendor": values[po_hmap["Vendor"] - 1] if "Vendor" in po_hmap else "",
                "customer": values[po_name_idx - 1] if po_name_idx and po_name_idx <= len(values) else "",
                "sheet": values[0],
                "vendorRoute": values[1],
            })
    except KeyError:
        po_row_count = 0
        po_rows = []

    # Warehouse short
    try:
        ws_short = find_sheet(wb, K.SHEET_WAREHOUSE_SHORT)
        wh_row_count = max(ws_short.max_row - 1, 0)
        known_wh = ["Shortages", "UNIT", "Product Name", "Code"]
        wh_hmap = header_map(ws_short, known_headers=known_wh)
        shortages: list[dict[str, Any]] = []
        wh_rows = []

        pname_col = None
        for hdr, idx in wh_hmap.items():
            if hdr and "product name" in str(hdr).lower():
                pname_col = idx
                break

        code_idx = wh_hmap.get("Code")
        unit_idx = wh_hmap.get("UNIT")
        bin_idx = wh_hmap.get("Bin")
        desc_idx = wh_hmap.get("Description")
        vendor_idx = wh_hmap.get("Vendor")
        qty_idx = wh_hmap.get("Qty")
        cust_idx = wh_hmap.get("Name")
        txn_idx = wh_hmap.get("Transaction Date")
        driver_idx = wh_hmap.get("Driver")
        qoh_idx = wh_hmap.get("Quantity On Hand")
        internal_bin_idx = (
            wh_hmap.get("BIN(Internal)") or wh_hmap.get("BIN (Internal)")
            or wh_hmap.get("internal bin") or wh_hmap.get("BIN")
        )

        def _val(values: tuple, idx: int | None) -> Any:
            if not idx or idx - 1 >= len(values):
                return ""
            v = values[idx - 1]
            return "" if v is None else v

        for row_idx, values in iter_data_rows(ws_short, known_headers=known_wh):
            shortage_val = values[0]
            if shortage_val is not None and shortage_val != "":
                shortages.append({
                    "code": str(values[code_idx-1]) if code_idx else "",
                    "productName": str(values[pname_col - 1]) if pname_col else "",
                    "unit": str(values[unit_idx-1]) if unit_idx else "CASE",
                    "shortage": float(shortage_val) if shortage_val else 0,
                })

            wh_rows.append({
                "id": f"{K.SHEET_WAREHOUSE_SHORT}:{row_idx}",
                "values": values,
                "shortage": values[0],
                "unit": values[1],
                "updateVendor": values[2],
                "productName": str(values[pname_col - 1]) if pname_col else "",
                "code": str(values[code_idx-1]) if code_idx else "",
                "bin": _val(values, bin_idx),
                "description": _val(values, desc_idx),
                "vendor": _val(values, vendor_idx),
                "qty": _val(values, qty_idx),
                "customer": _val(values, cust_idx),
                "transactionDate": str(_val(values, txn_idx) or ""),
                "driver": str(_val(values, driver_idx) or "").strip(),
                "qoh": _val(values, qoh_idx),
                "internalBin": _val(values, internal_bin_idx),
            })
    except KeyError:
        wh_row_count = 0
        shortages = []
        wh_rows = []

    # Drivers — build the ordered list for the UI.
    # Priority: if Driver Setup is already filled (re-hydrating a saved session),
    # use that sequence as-is.  Otherwise (fresh Part 1), apply the spec §6.2 /
    # §7.4.3 rule: default order filtered to names present in the data, then any
    # unlisted data-drivers appended alphabetically.
    try:
        ds = find_sheet(wb, K.SHEET_DRIVER_SETUP)
        drivers: list[str] = []
        for row in ds.iter_rows(min_row=2, values_only=True):
            val = row[0]
            if val not in (None, ""):
                drivers.append(str(val))
    except KeyError:
        drivers = []

    if not drivers:
        # Collect every distinct driver name that appears in the actual data.
        data_drivers: set[str] = set()
        for sheet_name in K.ROUTABLE_SHEETS:
            try:
                _ws = find_sheet(wb, sheet_name)
                from nightshift.sheet_utils import header_index
                d_col = header_index(_ws, "Driver")
                for _, vals in iter_data_rows(_ws, known_headers=["Driver"]):
                    v = vals[d_col - 1]
                    if v not in (None, ""):
                        data_drivers.add(str(v).strip())
            except (KeyError, ValueError):
                continue

        # 1. Default sequence, filtered to names present in the data.
        drivers = [d for d in K.DEFAULT_JETRO_DRIVER_ORDER if d in data_drivers]
        # 2. Extras not in the default list — appended alphabetically (spec §6.2).
        default_set = set(K.DEFAULT_JETRO_DRIVER_ORDER)
        extras = sorted(d for d in data_drivers if d not in default_set)
        drivers.extend(extras)

    # Vendors — collect from the PO sheet's *original* Vendor column.
    # After Part 1 reshapes the sheet, column B is "Vendor (route)" (empty
    # until the user picks a route) and the source Vendor column shifts
    # right by 3. We use header_map with an exact-name match so we don't
    # accidentally pick up "Vendor (route)" via a substring hit.
    vendors: list[str] = []
    try:
        po_ws = find_sheet(wb, K.SHEET_PO)
        po_h = header_map(po_ws)
        vendor_col = next(
            (idx for hdr, idx in po_h.items()
             if str(hdr).strip().lower() == "vendor"),
            None,
        )
        if vendor_col:
            seen: set[str] = set()
            for _, values in iter_data_rows(po_ws):
                if vendor_col - 1 >= len(values):
                    continue
                v = values[vendor_col - 1]
                if v in (None, ""):
                    continue
                name = str(v).replace(",", "").strip()
                if name and name not in seen:
                    vendors.append(name)
                    seen.add(name)
    except KeyError:
        pass

    return {
        "allOrders": {"rowCount": ao_row_count, "rows": ao_rows},
        "jetroSource": {"rowCount": js_row_count, "rows": js_rows},
        "po": {"rowCount": po_row_count, "rows": po_rows},
        "warehouseShort": {"rowCount": wh_row_count, "shortages": shortages, "rows": wh_rows},
        "drivers": drivers,
        "vendors": vendors,
    }


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------

@app.post("/part1")
async def run_part1(file: UploadFile = File(...)):
    """Upload an All-Orders xlsx, run Part 1 pipeline, return preview."""
    if not file.filename or not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are accepted.")

    job_id = str(uuid.uuid4())[:8]
    job_dir = _job_dir(job_id)

    # Save uploaded file
    contents = await file.read()
    src_path = job_dir / "upload.xlsx"
    src_path.write_bytes(contents)

    try:
        wb = load_workbook(BytesIO(contents))

        # Repair known header typos (e.g. "Vender" -> "Vendor") before any
        # downstream code does column lookups.
        normalize_header_typos(wb)

        # Run Part 1 pipeline
        compile_all_orders(wb)
        extract_jetro_and_po(wb)
        calculate_shortages(wb)
        setup_sheets_and_dropdowns(wb)

        # Save the Part 1 output
        out_path = job_dir / "part1.xlsx"
        wb.save(str(out_path))

        # Extract the preview
        preview = _extract_preview(wb)

    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Part 1 processing failed: {e}")

    return {"jobId": job_id, "status": "part1_complete", "preview": preview}


@app.get("/status/{job_id}")
async def get_status(job_id: str):
    job_dir = JOBS_DIR / job_id
    if not job_dir.exists():
        raise HTTPException(status_code=404, detail="Job not found.")
    has_part1 = (job_dir / "part1.xlsx").exists()
    return {"jobId": job_id, "status": "part1_complete" if has_part1 else "pending"}


# Output filenames are stable per job (Part 3 / Part 4 write to these paths).
# The Dry / Freezer / WH Pickup PDFs are emitted as separate files so each can
# be printed and handed out independently (operational request).
_OUTPUT_MAP: dict[str, dict[str, str]] = {
    "poReport":      {"xlsx": "po_report.xlsx", "pdf": "po_report.pdf"},
    "dryFreezerWh":  {
        "xlsx":        "pick_sheets.xlsx",
        "dryPdf":      "dry.pdf",
        "freezerPdf":  "freezer.pdf",
        "whPickupPdf": "wh_pickup.pdf",
    },
    "jetroWorkbook": {"xlsx": "jetro.xlsx"},
    "jetroPdf":      {"pdf": "jetro_report.pdf"},
}


def _collect_outputs(job_dir: Path) -> dict[str, dict[str, str]]:
    """Reconstruct the /part3-4 outputs payload from files on disk."""
    out: dict[str, dict[str, str]] = {}
    for key, files in _OUTPUT_MAP.items():
        present = {kind: name for kind, name in files.items()
                   if (job_dir / name).exists()}
        if present:
            out[key] = present
    return out


@app.get("/jobs/{job_id}")
async def get_job(job_id: str):
    """Rehydrate a job for a reloaded frontend.

    Returns the highest-stage status reachable from disk along with a freshly
    re-extracted preview (from part1.xlsx) and the outputs map (filenames of
    whichever Part 3 / Part 4 artifacts exist).
    """
    job_dir = JOBS_DIR / job_id
    part1_path = job_dir / "part1.xlsx"
    if not job_dir.exists() or not part1_path.exists():
        raise HTTPException(status_code=404, detail="Job not found.")

    outputs = _collect_outputs(job_dir)
    if outputs:
        status = "complete"
    elif (job_dir / "part2.xlsx").exists():
        status = "part2_complete"
    else:
        status = "part1_complete"

    # Preview is always rebuilt from part1.xlsx so the Review/Routing steps
    # have the same data they had on the original upload.
    wb = load_workbook(str(part1_path), data_only=True)
    try:
        normalize_header_typos(wb)
        preview = _extract_preview(wb)
    finally:
        wb.close()

    return {"jobId": job_id, "status": status, "preview": preview, "outputs": outputs}


@app.get("/files/{job_id}/{filename}")
async def download_file(job_id: str, filename: str):
    """Download a generated file for a given job."""
    job_dir = JOBS_DIR / job_id
    file_path = job_dir / filename
    if not file_path.exists():
        raise HTTPException(status_code=404, detail=f"File {filename} not found for job {job_id}.")
    return FileResponse(
        path=str(file_path),
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# Theme palette suffixes (alpha-prefixed RGB strings stored by openpyxl).
_THEME_BAND_SUFFIX = "1F2937"
_THEME_HEADER_SUFFIX = "4472C4"
_THEME_PRODUCE_SUFFIX = "D3D3D3"
_THEME_Z_DRIVER_SUFFIX = "B5BFC9"


def _row_kind(ws, row_idx: int) -> str:
    """Classify a row by the fill color of its first cell."""
    cell = ws.cell(row=row_idx, column=1)
    fill = getattr(cell, "fill", None)
    fg = getattr(fill, "fgColor", None) if fill else None
    rgb = getattr(fg, "rgb", None) if fg else None
    if not rgb or not isinstance(rgb, str):
        return "data"
    s = rgb.upper()
    if s.endswith(_THEME_BAND_SUFFIX):
        return "band"
    if s.endswith(_THEME_HEADER_SUFFIX):
        return "header"
    if s.endswith(_THEME_PRODUCE_SUFFIX):
        return "produce"
    if s.endswith(_THEME_Z_DRIVER_SUFFIX):
        return "z-driver"
    return "data"


@app.get("/preview/{job_id}/{filename}")
async def preview_workbook(job_id: str, filename: str):
    """Return every sheet of a generated workbook as JSON for a read-only grid.

    Each row is tagged with a `kind` ("band" / "header" / "produce" / "data")
    derived from the styled fill on column A, so the frontend can mirror the
    spreadsheet theme without shipping full per-cell style data.
    """
    job_dir = JOBS_DIR / job_id
    file_path = job_dir / filename
    if not filename.endswith(".xlsx") or not file_path.exists():
        raise HTTPException(status_code=404, detail=f"Workbook {filename} not found.")

    # Cap the rows returned per sheet so large reports stay responsive.
    MAX_ROWS = 2000
    # Styles are needed for row-kind detection, so we don't use read_only mode.
    wb = load_workbook(str(file_path), data_only=True)
    sheets: list[dict[str, Any]] = []
    for ws in wb.worksheets:
        rows: list[dict[str, Any]] = []
        truncated = False
        max_row = min(ws.max_row, MAX_ROWS)
        if ws.max_row > MAX_ROWS:
            truncated = True

        # Detect centered columns from the first non-band, non-header data-style
        # row (regular data, produce, or z-driver) so the SheetViewer renders
        # Qty/Price/etc. centered like the xlsx.
        center_cols: list[int] = []
        data_kinds = {"data", "produce", "z-driver"}
        for r_idx in range(1, max_row + 1):
            if _row_kind(ws, r_idx) not in data_kinds:
                continue
            cols_centered: list[int] = []
            for c_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                horiz = getattr(getattr(cell, "alignment", None), "horizontal", None)
                if horiz == "center":
                    cols_centered.append(c_idx)
            if cols_centered:
                center_cols = cols_centered
                break

        for r_idx in range(1, max_row + 1):
            cells = ["" if c.value is None else c.value
                     for c in ws[r_idx]]
            rows.append({"kind": _row_kind(ws, r_idx), "cells": cells})
        sheets.append({
            "name": ws.title,
            "rows": rows,
            "centerCols": center_cols,
            "truncated": truncated,
        })
    wb.close()
    return {"jobId": job_id, "filename": filename, "sheets": sheets}


@app.post("/part2")
async def run_part2(payload: dict[str, Any]):
    """Apply routing decisions."""
    job_id = payload.get("jobId")
    if not job_id:
        raise HTTPException(status_code=400, detail="Missing jobId.")

    job_dir = JOBS_DIR / job_id
    part1_path = job_dir / "part1.xlsx"
    if not part1_path.exists():
        raise HTTPException(status_code=404, detail="Part 1 output not found. Upload first.")

    try:
        wb = load_workbook(str(part1_path))
        normalize_header_typos(wb)

        sheet_routing = payload.get("sheetRoutingDecisions", [])
        ws_routing = payload.get("warehouseShortDecisions", [])
        driver_seq = payload.get("driverPullSequence", [])
        cell_edits = payload.get("cellEdits", [])

        apply_routing(wb, sheet_routing, ws_routing, driver_seq, cell_edits)
        
        out_path = job_dir / "part2.xlsx"
        wb.save(str(out_path))
        
        return {
            "jobId": job_id,
            "status": "part2_complete",
        }
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Part 2 processing failed: {e}")

@app.post("/part3-4")
async def run_part3_4(payload: dict[str, Any]):
    """Generate final reports."""
    job_id = payload.get("jobId")
    if not job_id:
        raise HTTPException(status_code=400, detail="Missing jobId.")

    job_dir = JOBS_DIR / job_id
    part2_path = job_dir / "part2.xlsx"
    if not part2_path.exists():
        raise HTTPException(status_code=404, detail="Part 2 output not found. Run Part 2 first.")

    try:
        wb = load_workbook(str(part2_path))
        normalize_header_typos(wb)

        # Run Part 3 & 4
        outputs3 = build_final_pick_sheets(wb, str(job_dir))
        outputs4 = build_jetro_branch(wb, str(job_dir))
        
        # Merge outputs
        all_outputs = {**outputs3, **outputs4}
        
        # Create a zip of all outputs (Spec §7.6)
        zip_path = job_dir / "all_outputs.zip"
        with zipfile.ZipFile(zip_path, "w") as zf:
            for category in all_outputs.values():
                for file_rel_path in category.values():
                    full_path = job_dir / file_rel_path
                    if full_path.exists():
                        zf.write(full_path, file_rel_path)
        
        return {
            "jobId": job_id,
            "status": "complete",
            "outputs": all_outputs
        }
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Part 3/4 processing failed: {e}")
