"""Synthetic raw-input workbook builder used by all tests.

The fixture exercises every branch of the spec:

  * numeric bin -> Jetro source (incl. bin == 0)
  * QOH == 0    -> PO
  * CASE / Unit / LBS shortage math
  * Shopping History trailing-U/C fuzzy match
  * blank QOH (instruction lines) skipped from shortages
  * Produce category rows for Jetro page grey shading
  * Multiple drivers for Freezer one / Freezer two split
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from nightshift import constants as K


# ---- All Orders -------------------------------------------------------------
# Header order (column A onward). We keep things in spec-defined order so the
# code matches what production workbooks look like.
ALL_ORDERS_HEADERS = [
    "Product Name", "Code", "Bin", "Description", "Vendor", "Address",
    "Qty", "Sell Price", "Unit Price Charged", "Total", "Customer Code",
    "Name", "Customer Address", "Date 1", "Order Date", "Date 3",
    "Transaction Date", "Driver", "Status", "Notes",
]

# (Product Name, Code, Bin, Description, Vendor, Qty, Name (customer), Driver, Category)
# Notes:
#  - Bin "5" is numeric -> Jetro
#  - Bin 0 (numeric zero) -> Jetro
#  - "DRY"/"FREEZER"/"COOLER"/"PICK UP" -> All Orders
#  - QOH=0 in inventory -> moves to PO after Jetro
#  - codes "1001U" appear in Shopping History as "1001" (trailing-U fuzzy)
ALL_ORDERS_ROWS = [
    # produce — Jetro (numeric bin 5)
    ("Roma Tomatoes",   "P-100",  "5", "Case 25 lb", "Jetro",  "10 Vine St",
     6, "Gengis Khan Torrance", "Glen"),
    # numeric bin 0 — also Jetro
    ("Iceberg Lettuce", "P-200",  0,   "Case",        "Jetro",  "10 Vine St",
     4, "Bennys Tacos Santa Monica", "Abraham"),
    # dry, normal stock
    ("Semolina Flour",  "1001U",  "DRY", "50 LBS bag",  "WH",   "55 Mill Rd",
     3, "Union Pizza Manhattan", "Glen"),
    # dry, shortage (LBS) — ordered 3 x 50lb avg = 150 lb; QOH 100 lb -> -50
    ("Beef Patty",      "BF1234", "DRY", "4oz frozen",  "WH",   "55 Mill Rd",
     3, "Mama's Hummus", "Glen"),
    ("Beef Patty",      "BF1234", "DRY", "4oz frozen",  "WH",   "55 Mill Rd",
     1, "Hummus Factory Manhattan", "Ramon"),
    # freezer, normal — drives Freezer page
    ("Churros Bavarian","FZ-77",  "FREEZER", "10in",    "WH",   "1 Cold Way",
     2, "Bennys Tacos Santa Monica", "Glen"),
    ("Pork Bellies",    "FZ-88",  "FREEZER", "Rind off",  "WH", "1 Cold Way",
     1, "CaCao Mexicatessen", "Abraham"),
    # cooler — will land on WH Pickup
    ("Milk 2%",         "C-50",   "COOLER", "Gallon",   "WH",   "Dock 1",
     2, "Mama's Hummus", "Gilberto"),
    # PO candidate — QOH=0 in inventory
    ("Olive Oil",       "OO-9",   "DRY", "1 gal",       "Sysco", "Sysco Dock",
     5, "Pollo Master", "Jose"),
    # CASE shortage — ordered 4, QOH 2 -> -2
    ("Fish Sauce",      "FS-12",  "DRY", "12/23oz",     "C Pacific Foods", "55 Mill Rd",
     4, "Gengis Khan Torrance", "Glen"),
    # blank-QOH instruction line — must be skipped from shortages
    ("Delivery Fee",    "FEE",    "DRY", "Fee line",    "WH",   "",
     1, "Mama's Hummus", "Glen"),
    # pick up
    ("Spring Water",    "SW-5",   "PICK UP", "Case",    "WH",   "",
     1, "CaCao Mexicatessen", "Adrian"),
]


def _set_row(ws, row_idx: int, mapping: dict[str, object]) -> None:
    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    for header, value in mapping.items():
        col = headers[header]
        ws.cell(row=row_idx, column=col, value=value)


def build_raw_workbook(path: Path) -> Path:
    """Write a fresh synthetic raw workbook to `path` and return the path."""
    wb = Workbook()

    ws = wb.active
    ws.title = K.SHEET_ALL_ORDERS
    ws.append(ALL_ORDERS_HEADERS)
    for (pname, code, bin_, desc, vend, addr, qty, name, driver) in ALL_ORDERS_ROWS:
        row = [
            pname, code, bin_, desc, vend, addr,
            qty, None, None, None, None,
            name, None, None, None, None,
            "2026-05-26", driver, None, None,
        ]
        ws.append(row)

    # item list (Web CSV) — provides Category Name / Cost / Selling / Unit per code
    il = wb.create_sheet(K.SHEET_ITEM_LIST)
    il.append(["SKU CODE", "Category Name", "Current Cost Price",
               "Current Selling Price", "Unit"])
    il.append(["P-100",  "Produce",  3.50, 5.00,  "CASE"])
    il.append(["P-200",  "Produce",  2.10, 3.25,  "CASE"])
    il.append(["1001U",  "Dry Goods", 18.0, 24.50, "CASE"])
    il.append(["BF1234", "Meat",      2.25, 3.10,  "LBS"])
    il.append(["FZ-77",  "Frozen",    9.0,  12.0,  "CASE"])
    il.append(["FZ-88",  "Frozen",    35.0, 48.0,  "LBS"])
    il.append(["C-50",   "Dairy",     2.50, 3.40,  "Unit"])
    il.append(["OO-9",   "Dry Goods", 22.0, 30.0,  "CASE"])
    il.append(["FS-12",  "Dry Goods", 11.0, 15.0,  "CASE"])
    il.append(["SW-5",   "Beverage",  6.0,  9.0,   "CASE"])
    il.append(["FEE",    "Service",   0,    0,     ""])  # blank-QOH companion

    # Inventory — Quantity On Hand drives PO + shortage math
    inv = wb.create_sheet(K.SHEET_INVENTORY)
    inv.append(["Item", "Quantity On Hand", "Cost",
                "Case Avg Weight", "BIN (Internal)"])
    inv.append(["P-100",  20, 3.5,  25, "PRD-1"])
    inv.append(["P-200",  15, 2.1,  20, "PRD-2"])
    inv.append(["1001U",  12, 18.0, 50, "DRY-1B"])
    inv.append(["BF1234", 100, 2.25, 50, "DRY-2C"])   # LBS shortage: 3*50 - 100 = +50? -> not short
    inv.append(["FZ-77",  10, 9.0,  10, "FREEZER-A11"])
    inv.append(["FZ-88",  50, 35.0, 45, "FREEZER-6B"])
    inv.append(["C-50",   5, 2.5,   1, "COOLER-1"])
    inv.append(["OO-9",   0, 22.0,  1, "DRY-3A"])     # QOH=0 -> PO
    inv.append(["FS-12",  2, 11.0,  1, "DRY-4A"])     # CASE shortage: 4-2 = -2
    inv.append(["SW-5",   30, 6.0,  1, "PICK-1"])
    inv.append(["FEE",    None, 0,  0, ""])           # blank QOH -> skip

    # Shopping History — "1001" (no trailing U) must still match All Orders "1001U"
    sh = wb.create_sheet(K.SHEET_SHOPPING_HISTORY)
    sh.append(["Item", "Product", "Bin", "Unit Price", "Case Price"])
    sh.append(["P-100",  "Roma Tomatoes Bulk",   "5",   0.40, 10.00])
    sh.append(["P-200",  "Iceberg Lettuce 24ct", "0",   0.30, 7.20])
    sh.append(["1001",   "Semolina/Gold Medal",  "1B",  0.45, 22.50])  # fuzzy match
    sh.append(["BF1234", "Beef Patty 4oz",       "2C",  0.06, 3.00])
    sh.append(["FZ-77",  "Churros 10in",         "A11", 0.95, 9.50])
    sh.append(["FZ-88",  "Pork Bellies",         "6B",  0.80, 36.00])
    sh.append(["C-50",   "Milk 2% Gal",          "1",   2.45, 2.45])
    sh.append(["OO-9",   "Olive Oil 1gal",       "3A",  4.50, 22.00])
    sh.append(["FS-12",  "Fish Sauce Lucky",     "4A",  0.95, 11.40])
    sh.append(["SW-5",   "Spring Water Case",    "1",   0.50, 6.00])

    wb.save(path)
    return path


if __name__ == "__main__":
    out = Path(__file__).parent / "_built_fixture.xlsx"
    build_raw_workbook(out)
    print(f"wrote {out}")
