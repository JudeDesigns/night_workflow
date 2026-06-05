"""Part 1 — Shortage math + Warehouse short builder (spec §3.4).

Sign convention: a shortage value is the **deficit**, stored as a negative
number. We compute it as `stock - demand` so:

  * CASE / Unit : shortage = QOH - sum(ordered Qty for code)
  * LBS         : shortage = QOH - sum(ordered Qty * CASE AVG WEIGHT)

A negative result means the warehouse is short by that amount, matching the
spec sentence "a negative result is a shortage" and the LBS PDF wording
("subtract that from Quantity On Hand"). The CASE/Unit sentence in the PDF
is loosely worded but the operationally meaningful form is identical.

Rows with a blank Quantity On Hand (instruction / fee lines) are skipped.
"""
from __future__ import annotations

from collections import defaultdict
from typing import Any

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from . import constants as K
from .codes import code_key
from .sheet_utils import find_sheet, header_index, header_map, iter_data_rows, find_header_row


def _to_number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip())
    except (ValueError, TypeError):
        return None


_LBS_ALIASES = {"LBS", "LB", "POUND", "POUNDS", "#"}
_CASE_ALIASES = {"CASE", "CASES", "CS", "CSE", "CTN", "CARTON"}
_UNIT_ALIASES = {"UNIT", "UNITS", "EA", "EACH", "PC", "PCS", "PIECE", "PIECES"}


def _classify_unit(unit_value: Any) -> str:
    """Map free-text unit values onto the canonical LBS/CASE/Unit buckets.

    The shortage formula branches on LBS (multiplied by CASE AVG WEIGHT) vs.
    everything else, so getting this wrong silently miscalculates the warehouse
    short. We accept the common variants seen in real data ("LB", "CASES",
    "EA", ...) rather than only the spec's canonical forms.
    """
    if unit_value is None:
        return ""
    s = str(unit_value).strip().upper().rstrip(".")
    if not s:
        return ""
    if s in _LBS_ALIASES:
        return K.UNIT_LBS
    if s in _CASE_ALIASES:
        return K.UNIT_CASE
    if s in _UNIT_ALIASES:
        return K.UNIT_UNIT
    return s


def calculate_shortages(wb: Workbook) -> list[dict[str, Any]]:
    """Compute shortages on All Orders, build Warehouse short, drop shorted rows.

    Returns a summary list (one dict per shorted code) for callers / previews.
    """
    src = find_sheet(wb, K.SHEET_ALL_ORDERS)
    known = ["Code", "Qty", "Quantity On Hand", "UNIT", "CASE AVG WEIGHT"]
    hmap = header_map(src, known_headers=known)
    code_col = hmap["Code"]
    qty_col = hmap["Qty"]
    qoh_col = hmap["Quantity On Hand"]
    unit_col = hmap["UNIT"]
    weight_col = hmap["CASE AVG WEIGHT"]

    # First pass: aggregate per code.
    by_code: dict[str, dict[str, Any]] = {}
    rows_by_code: dict[str, list[int]] = defaultdict(list)
    for row_idx, values in iter_data_rows(src, known_headers=known):
        key = code_key(values[code_col - 1])
        if not key:
            continue
        qty = _to_number(values[qty_col - 1]) or 0.0
        qoh = _to_number(values[qoh_col - 1])
        unit = _classify_unit(values[unit_col - 1])
        weight = _to_number(values[weight_col - 1]) or 0.0
        entry = by_code.setdefault(key, {
            "code": norm_first(values[code_col - 1]),
            "qoh": qoh, "unit": unit, "weight": weight, "ordered": 0.0,
        })
        entry["ordered"] += qty
        rows_by_code[key].append(row_idx)

    # Second pass: decide which codes are shorted.
    shorted: dict[str, float] = {}
    summary: list[dict[str, Any]] = []
    for key, entry in by_code.items():
        if entry["qoh"] is None:  # blank QOH -> instruction/fee row, skip
            continue
        if entry["unit"] == K.UNIT_LBS:
            shortage = entry["qoh"] - (entry["ordered"] * entry["weight"])
        else:  # CASE / Unit / unknown
            shortage = entry["qoh"] - entry["ordered"]
        if shortage < 0:
            shorted[key] = shortage
            summary.append({
                "code": entry["code"], "unit": entry["unit"] or "CASE",
                "shortage": shortage,
            })

    if not shorted:
        return summary

    # Build Warehouse short with the spec-mandated front columns A/B/C.
    ws_short = wb.create_sheet(K.SHEET_WAREHOUSE_SHORT)
    h_idx = find_header_row(src, *known)
    src_headers = [c.value for c in src[h_idx]]
    ws_short.append(["Shortages", "UNIT", "Update vendor", *src_headers])

    shorted_row_indices: list[int] = []
    for key in shorted:
        is_first = True
        for r_idx in rows_by_code[key]:
            row_values = [src.cell(row=r_idx, column=c).value
                          for c in range(1, src.max_column + 1)]
            shortage_value = shorted[key] if is_first else None
            ws_short.append([shortage_value, by_code[key]["unit"] or "CASE",
                             None, *row_values])
            is_first = False
            shorted_row_indices.append(r_idx)

    # Drop the shorted rows from All Orders (bottom-up to preserve indices).
    for r_idx in sorted(shorted_row_indices, reverse=True):
        src.delete_rows(r_idx, 1)

    return summary


def norm_first(value: Any) -> str:
    """Return the original code text (preserving its as-stored form)."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()
