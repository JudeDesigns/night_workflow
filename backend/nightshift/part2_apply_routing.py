"""Part 2 — Apply the Routing (spec §4)."""
from __future__ import annotations

from typing import Any

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from . import constants as K
from .codes import code_key
from .sheet_utils import find_sheet, header_index, header_map, int_if_whole


def _copy_row_to(ws_dst: Worksheet, values: list[Any]) -> int:
    """Append `values` to `ws_dst` and return the new row index."""
    new_row = ws_dst.max_row + 1
    for ci, v in enumerate(values, start=1):
        ws_dst.cell(row=new_row, column=ci, value=v)
    return new_row


def _parse_decision_id(raw_id: Any) -> tuple[str, int] | None:
    """Split a frontend row id ('<sheet>:<row>') into (sheet_name, row_idx)."""
    if not isinstance(raw_id, str) or ":" not in raw_id:
        return None
    sheet_name, row_str = raw_id.rsplit(":", 1)
    try:
        return sheet_name, int(row_str)
    except ValueError:
        return None


# Frontend field name -> ordered list of header names to try in the target sheet.
# The first header that exists in the sheet's header_map wins. This lets the
# same field map to slightly different column titles across sheets (e.g. the
# internal bin is "BIN(Internal)" on All Orders / PO but doesn't exist on
# Jetro Source — there "bin" should fall through to the numeric "Bin").
_FIELD_HEADERS: dict[str, list[str]] = {
    "productName":  ["Product Name"],
    "code":         ["Code"],
    "bin":          ["BIN(Internal)", "BIN (Internal)", "internal bin", "Bin"],
    "internalBin":  ["BIN(Internal)", "BIN (Internal)", "internal bin", "BIN"],
    "vendor":       ["Vendor"],
    "description":  ["Description"],
    "qty":          ["Qty"],
    "qoh":          ["Quantity On Hand"],
    "customer":     ["Name"],
    "driver":       ["Driver"],
    "shortage":     ["Shortages"],
    "unit":         ["UNIT"],
}

# Jetro Source displays the numeric Bin (col F), so an edit to "bin" there
# should land on the numeric column instead of the internal-bin alias chain.
_FIELD_HEADERS_BY_SHEET: dict[str, dict[str, list[str]]] = {
    K.SHEET_JETRO_SOURCE: {"bin": ["Bin"]},
}

# Numeric fields — convert input strings to float before writing so downstream
# sum/sort logic still works.
_NUMERIC_FIELDS = {"qty", "qoh", "shortage"}


def _coerce(field: str, value: Any) -> Any:
    """Cast incoming JSON value to the type the workbook expects."""
    # Treat blank / whitespace-only input as "cleared".
    if value is None or (isinstance(value, str) and not value.strip()):
        return None if field in _NUMERIC_FIELDS else ""
    if field in _NUMERIC_FIELDS:
        try:
            return float(value)
        except (TypeError, ValueError):
            return value
    return value


def apply_cell_edits(wb: Workbook, cell_edits: list[dict[str, Any]]) -> None:
    """Write user edits from the Routing UI back to the workbook.

    Each edit is `{id: "<sheet>:<row>", field: "<fieldName>", value: <any>}`.
    Edits run **before** any routing logic so the rest of Part 2 (sheet moves,
    WS re-routing, PO totals) operates on the corrected data.
    """
    for edit in cell_edits:
        parsed = _parse_decision_id(edit.get("id"))
        if not parsed:
            continue
        sheet_name, row_idx = parsed
        field = edit.get("field")
        if not field or field not in _FIELD_HEADERS:
            continue

        try:
            ws = find_sheet(wb, sheet_name)
        except KeyError:
            continue

        hmap = header_map(ws)
        # Sheet-specific override first, then the general fallback chain.
        candidates = (
            _FIELD_HEADERS_BY_SHEET.get(ws.title, {}).get(field)
            or _FIELD_HEADERS[field]
        )
        col_idx = next((hmap[h] for h in candidates if h in hmap), None)
        if not col_idx:
            continue

        # openpyxl's ws.cell(value=None) is a no-op, so assign via .value
        # directly to support clearing a cell.
        ws.cell(row=row_idx, column=col_idx).value = _coerce(field, edit.get("value"))


def apply_sheet_dropdowns(wb: Workbook, routing_decisions: list[dict[str, Any]]) -> None:
    """Apply the Sheet and Vendor dropdown routing for All Orders, Jetro, PO."""
    for decision in routing_decisions:
        parsed = _parse_decision_id(decision.get("id"))
        if not parsed:
            continue
        sheet_name, row_idx = parsed
        new_sheet = decision.get("sheet")
        vendor = decision.get("vendor")

        try:
            ws = find_sheet(wb, sheet_name)
        except KeyError:
            continue
        if new_sheet:
            ws.cell(row=row_idx, column=1, value=new_sheet)
        if vendor not in (None, ""):
            ws.cell(row=row_idx, column=2, value=vendor)

    moves: list[tuple[Worksheet, Worksheet, int, list[Any]]] = []
    
    for sheet_name in K.ROUTABLE_SHEETS:
        try:
            ws = find_sheet(wb, sheet_name)
        except KeyError:
            continue
            
        for r in range(ws.max_row, 1, -1):
            row_values = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
            if all(v in (None, "") for v in row_values[3:]):
                continue
                
            target_sheet_name = str(row_values[0]).strip()
            if target_sheet_name.lower() != sheet_name.lower() and target_sheet_name.lower() in [s.lower() for s in K.ROUTABLE_SHEETS]:
                try:
                    target_ws = find_sheet(wb, target_sheet_name)
                    row_values[0] = target_ws.title
                    moves.append((ws, target_ws, r, row_values))
                except KeyError:
                    pass

    for src_ws, tgt_ws, row_idx, values in moves:
        src_ws.delete_rows(row_idx, 1)
        _copy_row_to(tgt_ws, values)


def apply_warehouse_short_routing(wb: Workbook, ws_decisions: list[dict[str, Any]]) -> None:
    """Apply routing from Warehouse short sheet based on UI decisions."""
    try:
        ws_short = find_sheet(wb, K.SHEET_WAREHOUSE_SHORT)
    except KeyError:
        return

    for decision in ws_decisions:
        parsed = _parse_decision_id(decision.get("id"))
        if not parsed:
            continue
        _, r = parsed
        update_vendor = decision.get("updateVendor")
        if r and update_vendor:
            ws_short.cell(row=r, column=3, value=update_vendor)

    moves: list[tuple[Worksheet, int, list[Any], str, str | None]] = []
    hmap = header_map(ws_short)
    code_col = hmap.get("Code")
    shortage_col = hmap.get("Shortages", 1)
    ws_qty_col = hmap.get("Qty")
    
    for r in range(ws_short.max_row, 1, -1):
        row_values = [ws_short.cell(row=r, column=c).value for c in range(1, ws_short.max_column + 1)]
        if all(v in (None, "") for v in row_values[3:]):
            continue
            
        update_vendor = str(row_values[2]).strip() if row_values[2] else ""
        if not update_vendor:
            continue

        target_ws_name = None
        new_bin = None
        new_vendor = None
        
        if update_vendor.lower() == K.WH_ROUTING_JETRO.lower():
            target_ws_name = K.SHEET_JETRO_SOURCE
        elif update_vendor.lower() in [b.lower() for b in K.ALL_ORDERS_BINS]:
            target_ws_name = K.SHEET_ALL_ORDERS
            new_bin = update_vendor
        elif update_vendor.lower() == K.WH_ROUTING_WH.lower():
            # Route to All Orders (not PO) so the item appears in the
            # WH Pickup report. The row's original Bin and Vendor are
            # preserved — no override needed.
            target_ws_name = K.SHEET_ALL_ORDERS
        else:
            target_ws_name = K.SHEET_PO
            new_vendor = update_vendor

        moves.append((ws_short, r, row_values, target_ws_name, new_vendor))
        
        if new_bin:
            try:
                bin_idx = hmap.get("Bin")
                if bin_idx:
                    row_values[bin_idx - 1] = new_bin
            except Exception:
                pass

    for src_ws, row_idx, values, tgt_name, new_vendor in moves:
        shortage_val = values[shortage_col - 1]
        code_val = values[code_col - 1] if code_col else None
        
        src_ws.delete_rows(row_idx, 1)
        
        if shortage_val not in (None, "") and code_val:
            key = code_key(code_val)
            for remain_r in range(2, src_ws.max_row + 1):
                c_val = src_ws.cell(row=remain_r, column=code_col).value
                if code_key(c_val) == key:
                    src_ws.cell(row=remain_r, column=shortage_col, value=shortage_val)
                    break
        
        # Build the destination row.
        # Warehouse short layout:  [Shortages, UNIT, Update vendor,
        #                           Product Name, Code, Bin, Description,
        #                           Vendor, Address, Qty, ...]
        # Target layout (PO / All Orders / Jetro source):
        #                          [Sheet, Vendor (route), Qty-or-Total-Qty,
        #                           Product Name, Code, Bin, ...]
        # So we keep WS cols 4+ (the standard order columns) and prepend the
        # 3 target-specific front cols. For PO, col 3 (Total Qty) is left
        # blank — recalculate_po_totals fills it in. For All Orders / Jetro
        # source, col 3 mirrors the row's ordered Qty per spec §3.5.
        try:
            tgt_ws = find_sheet(wb, tgt_name)
        except KeyError:
            continue

        front_qty: Any = None
        if tgt_ws.title.strip().lower() in (
            K.SHEET_ALL_ORDERS.lower(), K.SHEET_JETRO_SOURCE.lower(),
        ) and ws_qty_col and ws_qty_col - 1 < len(values):
            front_qty = values[ws_qty_col - 1]

        front_vendor = new_vendor if new_vendor else ""
        insert_values = [tgt_ws.title, front_vendor, front_qty] + list(values[3:])
        _copy_row_to(tgt_ws, insert_values)


def recalculate_po_totals(wb: Workbook) -> None:
    """Recalculate PO column C (Total Qty) after all routing is done."""
    try:
        po_ws = find_sheet(wb, K.SHEET_PO)
    except KeyError:
        return
        
    hmap = header_map(po_ws)
    code_col = hmap.get("Code")
    qty_col = hmap.get("Qty")
    total_qty_col = hmap.get("Total Qty", 3)
    
    if not (code_col and qty_col):
        return

    totals: dict[str, float] = {}
    for r in range(2, po_ws.max_row + 1):
        key = code_key(po_ws.cell(row=r, column=code_col).value)
        if not key:
            continue
        q = po_ws.cell(row=r, column=qty_col).value
        try:
            val = float(q) if q not in (None, "") else 0.0
            totals[key] = totals.get(key, 0.0) + val
        except (TypeError, ValueError):
            pass

    seen: set[str] = set()
    for r in range(2, po_ws.max_row + 1):
        key = code_key(po_ws.cell(row=r, column=code_col).value)
        if not key:
            continue
        if key not in seen:
            po_ws.cell(row=r, column=total_qty_col, value=int_if_whole(totals.get(key, 0.0)))
            seen.add(key)
        else:
            po_ws.cell(row=r, column=total_qty_col, value=None)


def drop_source_sheets(wb: Workbook) -> None:
    """Drop the raw source sheets no longer needed."""
    for sname in K.SOURCE_SHEETS:
        try:
            ws = find_sheet(wb, sname)
            wb.remove(ws)
        except KeyError:
            pass


def apply_driver_sequence(wb: Workbook, driver_sequence: list[str]) -> None:
    """Store the driver sequence to the Driver Setup sheet."""
    try:
        ws = find_sheet(wb, K.SHEET_DRIVER_SETUP)
        ws.delete_rows(2, max(ws.max_row, 2))
        for i, driver in enumerate(driver_sequence, start=2):
            fg = "Freezer one" if i <= 4 else "Freezer two"
            ws.cell(row=i, column=1, value=driver)
            ws.cell(row=i, column=2, value=fg)
    except KeyError:
        pass


def apply_routing(
    wb: Workbook,
    sheet_decisions: list[dict[str, Any]],
    ws_decisions: list[dict[str, Any]],
    driver_sequence: list[str],
    cell_edits: list[dict[str, Any]] | None = None,
) -> None:
    """Run the entire Part 2 pipeline."""
    # Cell edits run first so subsequent routing reads the corrected values.
    if cell_edits:
        apply_cell_edits(wb, cell_edits)
    if driver_sequence:
        apply_driver_sequence(wb, driver_sequence)
    apply_sheet_dropdowns(wb, sheet_decisions)
    apply_warehouse_short_routing(wb, ws_decisions)
    recalculate_po_totals(wb)
    drop_source_sheets(wb)
