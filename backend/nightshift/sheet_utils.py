"""Low-level openpyxl helpers — header lookup, row reads, dropdown writes."""
from __future__ import annotations

from typing import Any, Iterable

from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet


def find_sheet(wb, name: str) -> Worksheet:
    """Locate a sheet by name, case-insensitive, ignoring extra whitespace.

    Matches if:
    1. Exact match (case-insensitive)
    2. Expected name is a substring of the actual title (e.g. "All Orders" -> "All orders Sheet.")
    3. Actual title is a substring of the expected name (e.g. "item list (Web CSV)" -> "item list")
    
    Raises KeyError if nothing matches.
    """
    target = name.strip().lower()
    
    # 1. Try exact match first
    for ws in wb.worksheets:
        title = ws.title.strip().lower()
        if title == target:
            return ws
            
    # 2. Try substring match (either way)
    for ws in wb.worksheets:
        title = ws.title.strip().lower()
        if target in title or title in target:
            return ws
            
    raise KeyError(f"Sheet not found: {name!r}; have {[s.title for s in wb.worksheets]}")


def find_header_row(ws: Worksheet, *known_headers: str) -> int:
    """Find the first row that likely contains the headers.
    
    Scans the first 500 rows. A row is considered a header row if it contains 
    at least TWO of the `known_headers` (case-insensitive substring match)
    AND does not contain metadata noise like 'Ref Number'.
    """
    if not known_headers:
        for r_idx in range(1, min(ws.max_row + 1, 501)):
            if any(cell.value is not None and str(cell.value).strip() != "" for cell in ws[r_idx]):
                return r_idx
        return 1
        
    normalized_known = [h.strip().lower() for h in known_headers]
    # We use very specific noise strings to avoid false positives on valid data
    # such as the 'All Orders' sheet name.
    metadata_noise = ["ref number:", "total delivery fee", "grand total"]
    
    for r_idx in range(1, min(ws.max_row + 1, 501)):
        row_values = []
        for cell in ws[r_idx]:
            if cell.value is not None:
                row_values.append(str(cell.value).strip().lower())
        
        # Skip if the row contains metadata noise in the first few columns
        if any(noise in " ".join(row_values[:3]) for noise in metadata_noise):
            continue

        match_count = 0
        found_knowns = set()
        for val in row_values:
            for known in normalized_known:
                if known in val and known not in found_knowns:
                    match_count += 1
                    found_knowns.add(known)
        
        # Require at least 2 matches (or all if less than 2 provided)
        required_matches = min(2, len(normalized_known))
        if match_count >= required_matches:
            # Final sanity check: Does it look like a header?
            # A header usually doesn't have 100+ characters in a single cell
            if any(len(val) > 100 for val in row_values):
                continue
            return r_idx
            
    return 1


def header_index(ws: Worksheet, *aliases: str) -> int:
    """Return the 1-based column index of the first matching header.

    Compares case-insensitively after stripping whitespace.
    Also handles substring matches (e.g. "Product Name" matches "Product Name (Full)").
    Scans for the header row first.
    Raises KeyError if no alias matches.
    """
    header_row_idx = find_header_row(ws, *aliases)
    normalized_aliases = [a.strip().lower() for a in aliases]
    
    # 1. Try exact matches first
    for cell in ws[header_row_idx]:
        if cell.value is None:
            continue
        val = str(cell.value).strip().lower()
        if val in normalized_aliases:
            return cell.column
            
    # 2. Try substring matches (either way)
    for cell in ws[header_row_idx]:
        if cell.value is None:
            continue
        val = str(cell.value).strip().lower()
        for alias in normalized_aliases:
            if alias in val or val in alias:
                return cell.column
                
    raise KeyError(f"Header not found: any of {aliases!r} on row {header_row_idx}; have {[c.value for c in ws[header_row_idx]]}")


def header_map(ws: Worksheet, known_headers: list[str] | None = None) -> dict[str, int]:
    """Map of header text (as-is) -> 1-based column index.
    
    If known_headers is provided, uses them to find the correct header row.
    """
    header_row_idx = find_header_row(ws, *(known_headers or []))
                
    out: dict[str, int] = {}
    for cell in ws[header_row_idx]:
        if cell.value is None:
            continue
        out[str(cell.value).strip()] = cell.column
    return out


def iter_data_rows(ws: Worksheet, known_headers: list[str] | None = None) -> Iterable[tuple[int, tuple]]:
    """Yield (row_number, values_tuple) for every data row.
    
    Starts after the detected header row.
    Stops at the first fully-empty row.
    """
    header_row_idx = find_header_row(ws, *(known_headers or []))
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=False):
        # A row is data only if it has at least some meaningful content
        # We skip rows that are completely empty or contain only whitespace
        if all(c.value in (None, "") or str(c.value).strip() == "" for c in row):
            continue
            
        # EXTRA ROBUSTNESS: Skip rows that look like noise/metadata
        # Extract only meaningful values (non-empty strings)
        meaningful_vals = [str(c.value).strip() for c in row if c.value is not None and str(c.value).strip() != ""]
        
        # If the row has very few meaningful columns populated, it's likely noise or a spacer
        # A valid order row should have at least: Product Name, Code, Bin, and Qty.
        if len(meaningful_vals) < 4: 
            continue
            
        # Specific check for metadata noise
        # We avoid "all order" here because it's a valid sheet name/dropdown value.
        if any(noise in meaningful_vals[0].lower() for noise in ["ref number:", "delivery fee:"]):
            continue

        # Skip if the row repeats headers (common in some report exports)
        if known_headers:
            match_count = 0
            for val in meaningful_vals:
                if any(h.lower() in val.lower() for h in known_headers):
                    match_count += 1
            if match_count >= len(known_headers) - 1 and len(known_headers) > 1:
                continue

        yield row[0].row, tuple(c.value for c in row)


def row_values(ws: Worksheet, row_idx: int) -> list[Any]:
    return [c.value for c in ws[row_idx]]


def int_if_whole(value: Any) -> Any:
    """Return int(value) when value is an integer-valued number, else value.

    Hides the trailing ``.0`` on aggregated totals (Qty / Total Qty) that
    happen to be whole numbers. Strings and non-numeric values pass through
    unchanged.
    """
    if value is None or value == "":
        return value
    try:
        f = float(value)
    except (TypeError, ValueError):
        return value
    if f == int(f):
        return int(f)
    return f


# Map of header typos found in real-world source files -> canonical spelling.
# Applied once at the start of Part 1 so every downstream lookup that asks for
# the canonical name finds the column regardless of how it's spelled.
HEADER_TYPO_MAP: dict[str, str] = {
    "vender": "Vendor",
}


def normalize_header_typos(wb) -> None:
    """Rewrite known mis-spelled headers (e.g. 'Vender' -> 'Vendor') in place.

    Scans the detected header row of every sheet. Idempotent.
    """
    for ws in wb.worksheets:
        h_idx = find_header_row(ws)
        for cell in ws[h_idx]:
            if cell.value is None:
                continue
            key = str(cell.value).strip().lower()
            if key in HEADER_TYPO_MAP:
                cell.value = HEADER_TYPO_MAP[key]


def set_header(ws: Worksheet, col_letter: str, header: str) -> None:
    ws[f"{col_letter}1"] = header


def col_letter(idx: int) -> str:
    return get_column_letter(idx)


def col_index(letter: str) -> int:
    return column_index_from_string(letter)


def add_list_validation(
    ws: Worksheet,
    col_letter_str: str,
    options: Iterable[str],
    first_data_row: int = 2,
    last_data_row: int | None = None,
) -> None:
    """Attach a dropdown list to a column range.

    Options are quoted as a comma-separated formula1 string. If the joined
    string would exceed Excel's 255-char formula1 limit, we fall back to
    writing the options into a hidden helper sheet and referencing the range.
    """
    opts = [str(o).replace(",", "") for o in options]  # commas removed per spec
    opts = [o for o in opts if o != ""]
    if not opts:
        return
    last = last_data_row if last_data_row is not None else max(ws.max_row, first_data_row)
    rng = f"{col_letter_str}{first_data_row}:{col_letter_str}{last}"

    joined = ",".join(opts)
    if len(joined) <= 250:
        dv = DataValidation(
            type="list",
            formula1=f'"{joined}"',
            allow_blank=True,
            showDropDown=False,
        )
        dv.add(rng)
        ws.add_data_validation(dv)
        return

    # Fallback: stash options on a hidden _dropdowns sheet.
    wb = ws.parent
    helper_name = "_dropdowns"
    helper = wb[helper_name] if helper_name in wb.sheetnames else wb.create_sheet(helper_name)
    helper.sheet_state = "hidden"
    start_col = helper.max_column + 1 if helper.max_row > 0 else 1
    helper_letter = get_column_letter(start_col)
    helper.cell(row=1, column=start_col, value=f"{ws.title}!{col_letter_str}")
    for i, o in enumerate(opts, start=2):
        helper.cell(row=i, column=start_col, value=o)
    ref = f"{helper_name}!${helper_letter}$2:${helper_letter}${len(opts) + 1}"
    dv = DataValidation(type="list", formula1=ref, allow_blank=True, showDropDown=False)
    dv.add(rng)
    ws.add_data_validation(dv)


def apply_number_format(ws: Worksheet, col_idx: int, fmt: str, first_data_row: int = 2) -> None:
    for r in range(first_data_row, ws.max_row + 1):
        ws.cell(row=r, column=col_idx).number_format = fmt
