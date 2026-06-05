"""Part 1 — Sheet layout, dropdowns, Driver Setup (spec §3.5)."""
from __future__ import annotations

from collections import defaultdict
from typing import Iterable

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from . import constants as K
from .codes import code_key
from .sheet_utils import (
    add_list_validation,
    col_letter,
    find_sheet,
    header_index,
    header_map,
    iter_data_rows,
    find_header_row,
)


# ---------------------------------------------------------------------------
# Vendor + driver collection
# ---------------------------------------------------------------------------

def _collect_vendors(wb: Workbook) -> list[str]:
    """Every distinct vendor name from the PO sheet (commas removed)."""
    try:
        ws = find_sheet(wb, K.SHEET_PO)
        col = header_index(ws, "Vendor")
    except KeyError:
        return []
    seen: dict[str, None] = {}
    for _, values in iter_data_rows(ws, known_headers=["Vendor"]):
        v = values[col - 1]
        if v in (None, ""):
            continue
        name = str(v).replace(",", "").strip()
        if name and name not in seen:
            seen[name] = None
    return list(seen)


def _collect_drivers(wb: Workbook) -> list[str]:
    """Every distinct driver name across the working sheets."""
    seen: dict[str, None] = {}
    for sheet_name in K.ROUTABLE_SHEETS:
        try:
            ws = find_sheet(wb, sheet_name)
            col = header_index(ws, "Driver")
        except KeyError:
            continue
        for _, values in iter_data_rows(ws, known_headers=["Driver"]):
            v = values[col - 1]
            if v in (None, ""):
                continue
            name = str(v).strip()
            if name and name not in seen:
                seen[name] = None
    return list(seen)


# ---------------------------------------------------------------------------
# Front-column shaping
# ---------------------------------------------------------------------------

def _prepend_front_columns(ws: Worksheet, qty_header: str) -> None:
    """Insert two new columns A and B and populate headers.

    Column A becomes the 'Sheet' dropdown; B becomes 'Vendor (route)';
    column C is later populated by the caller with `qty_header` ('Qty' or
    'Total Qty'). All existing data shifts right.
    """
    # We need to insert columns on the detected header row, but openpyxl
    # insert_cols inserts for the whole sheet. We'll need to move the headers
    # if they weren't on row 1.
    h_idx = find_header_row(ws, "Code", "Qty", "Product Name")
    
    ws.insert_cols(1, amount=3)
    ws.cell(row=h_idx, column=1, value="Sheet")
    ws.cell(row=h_idx, column=2, value="Vendor (route)")
    ws.cell(row=h_idx, column=3, value=qty_header)


def _populate_front_columns(
    ws: Worksheet,
    own_sheet: str,
    qty_total_by_code: dict[str, float] | None = None,
) -> None:
    """Fill columns A/B/C on data rows.

    - Column A defaults to the row's own sheet name.
    - Column B defaults to the row's own Vendor.
    - Column C either mirrors ordered Qty (when qty_total_by_code is None)
      or holds the per-code Total Qty on the first row of each code.
    """
    hmap = header_map(ws, known_headers=["Code", "Qty", "Vendor", "BIN(Internal)", "BIN (Internal)", "internal bin", "BIN"])
    code_col = hmap.get("Code")
    qty_col = hmap.get("Qty")
    vendor_col = hmap.get("Vendor")
    # Internal Bin for frontend routing preview
    internal_bin_col = hmap.get("BIN(Internal)") or hmap.get("BIN (Internal)") or hmap.get("internal bin") or hmap.get("BIN")
    seen_codes: set[str] = set()
    
    for row_idx, values in iter_data_rows(ws, known_headers=["Code", "Qty", "Vendor"]):
        ws.cell(row=row_idx, column=1, value=own_sheet)
        if vendor_col:
            vname = values[vendor_col - 1]
            ws.cell(row=row_idx, column=2, value=
                    str(vname).replace(",", "").strip() if vname not in (None, "") else None)
        if qty_total_by_code is not None and code_col:
            key = code_key(values[code_col - 1])
            if key and key not in seen_codes:
                ws.cell(row=row_idx, column=3, value=qty_total_by_code.get(key))
                seen_codes.add(key)
            else:
                ws.cell(row=row_idx, column=3, value=None)
        elif qty_col:
            ws.cell(row=row_idx, column=3, value=values[qty_col - 1])


def _po_totals(ws: Worksheet) -> dict[str, float]:
    hmap = header_map(ws, known_headers=["Code", "Qty"])
    code_col, qty_col = hmap.get("Code"), hmap.get("Qty")
    totals: dict[str, float] = defaultdict(float)
    if not (code_col and qty_col):
        return {}
    for _, values in iter_data_rows(ws, known_headers=["Code", "Qty"]):
        key = code_key(values[code_col - 1])
        q = values[qty_col - 1]
        if not key:
            continue
        try:
            totals[key] += float(q) if q not in (None, "") else 0.0
        except (TypeError, ValueError):
            pass
    return dict(totals)


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def setup_sheets_and_dropdowns(wb: Workbook) -> None:
    """Spec §3.5 — front columns, dropdowns, sort, Driver Setup sheet."""
    vendors = _collect_vendors(wb)
    drivers = _collect_drivers(wb)
    vendor_route_options = [*vendors, K.WH_ROUTING_JETRO]
    sheet_options = ["All orders", K.SHEET_JETRO_SOURCE, K.SHEET_PO]
    wh_short_options = [K.WH_ROUTING_JETRO, *K.ALL_ORDERS_BINS, *vendors]

    # PO is shaped slightly differently: column C is Total Qty.
    for sheet_name in K.ROUTABLE_SHEETS:
        ws = find_sheet(wb, sheet_name)
        if sheet_name == K.SHEET_PO:
            totals = _po_totals(ws)
            _prepend_front_columns(ws, "Total Qty")
            _populate_front_columns(ws, K.SHEET_PO, qty_total_by_code=totals)
        else:
            _prepend_front_columns(ws, "Qty")
            _populate_front_columns(ws, sheet_name)

        last = ws.max_row
        add_list_validation(ws, "A", sheet_options, last_data_row=last)
        add_list_validation(ws, "B", vendor_route_options, last_data_row=last)

    # Sort All Orders by Product Name (A->Z).
    _sort_by_product_name(find_sheet(wb, K.SHEET_ALL_ORDERS))

    # Warehouse short dropdown.
    try:
        ws_short = find_sheet(wb, K.SHEET_WAREHOUSE_SHORT)
        add_list_validation(ws_short, "C", wh_short_options,
                            last_data_row=ws_short.max_row)
    except KeyError:
        pass  # no shortages

    _build_driver_setup_sheet(wb, drivers)


def _sort_by_product_name(ws: Worksheet) -> None:
    h_idx = find_header_row(ws, "Product Name")
    hmap = header_map(ws, known_headers=["Product Name"])
    name_col = hmap.get("Product Name")
    if not name_col or ws.max_row <= h_idx:
        return
    headers = [ws.cell(row=h_idx, column=c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for row_idx, values in iter_data_rows(ws, known_headers=["Product Name"]):
        rows.append(list(values))
    rows.sort(key=lambda row: (str(row[name_col - 1] or "")).lower())
    
    ws.delete_rows(h_idx, ws.max_row)
    ws.append(headers)
    for r in rows:
        ws.append(r)


def _build_driver_setup_sheet(wb: Workbook, drivers: Iterable[str]) -> None:
    ws = wb.create_sheet(K.SHEET_DRIVER_SETUP)
    ws.append(["Order", "Freezer group"])
    # Leave the Order column blank for the office to fill in; preload the
    # Freezer group formula so it auto-fills as the office types.
    max_drivers = max(8, len(list(drivers)) + 2)
    for i in range(2, max_drivers + 2):
        ws.cell(row=i, column=2, value=f'=IF(A{i}="","",IF(ROW()-1<=3,"Freezer one","Freezer two"))')
    add_list_validation(ws, "A", list(drivers), last_data_row=max_drivers + 1)
