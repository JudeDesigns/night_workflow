"""Part 1 — Get the Results (spec §3).

Pipeline:
    1. Compile  : merge item list / Inventory / Shopping History into All Orders
    2. Extract  : numeric-bin -> Jetro source ; QOH=0 -> PO ; rest -> All Orders
    3. Shortages: per-code math, build Warehouse short, drop shorted rows
    4. Set up   : Sheet + Vendor(route) dropdowns, front Qty / PO Total Qty,
                  Warehouse short Update vendor dropdown, sort by Product Name,
                  empty Driver Setup sheet.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from . import constants as K
from .codes import code_key, code_key_loose, norm_code
from .sheet_utils import (
    add_list_validation,
    apply_number_format,
    col_letter,
    find_sheet,
    header_index,
    header_map,
    iter_data_rows,
    find_header_row,
)


def _is_numeric_bin(value: Any) -> bool:
    """A numeric bin (including 0) — used to identify Jetro rows."""
    if value is None or value == "":
        return False
    if isinstance(value, bool):
        return False
    if isinstance(value, (int, float)):
        return True
    s = str(value).strip()
    if not s:
        return False
    try:
        float(s)
        return True
    except ValueError:
        return False


def _build_compile_index(ws: Worksheet, code_header: str, loose: bool) -> dict[str, dict[str, Any]]:
    """Return {normalized_code: {header: value}} for every row in `ws`.

    `loose=True` strips a trailing U/C for Shopping History matching.
    """
    hmap = header_map(ws, known_headers=[code_header])
    if code_header not in hmap:
        # Fallback to loose check if exact header not found in hmap keys
        # because header_map might have slightly different casing/whitespace
        found_key = None
        for k in hmap.keys():
            if code_header.lower() in k.lower():
                found_key = k
                break
        if not found_key:
            raise KeyError(
                f"Sheet {ws.title!r} missing code column {code_header!r}; "
                f"have {list(hmap)}"
            )
        code_col = hmap[found_key]
    else:
        code_col = hmap[code_header]
        
    out: dict[str, dict[str, Any]] = {}
    for row_idx, values in iter_data_rows(ws, known_headers=[code_header]):
        raw = values[code_col - 1]
        key = code_key_loose(raw) if loose else code_key(raw)
        if not key:
            continue
        # First occurrence wins (matches Excel lookup behaviour).
        if key in out:
            continue
        out[key] = {h: values[ci - 1] for h, ci in hmap.items()}
    return out


def _norm_header(h: str) -> str:
    """Collapse whitespace and lowercase a header for tolerant matching."""
    return "".join(str(h).split()).lower()


def _record_get(record: dict[str, Any], src_header: str) -> Any:
    """Fetch a value from a compile record, tolerant of header spacing/case.

    Source sheets sometimes spell a header differently from the spec
    (e.g. ``BIN(Internal)`` vs ``BIN (Internal)``). Match on a normalized
    key so the lookup still succeeds.
    """
    if src_header in record:
        return record[src_header]
    target = _norm_header(src_header)
    for k, v in record.items():
        if _norm_header(k) == target:
            return v
    return None


def compile_all_orders(wb: Workbook) -> None:
    """Spec §3.1 — copy mapped columns from source sheets into All Orders."""
    all_orders = find_sheet(wb, K.SHEET_ALL_ORDERS)
    item_list = find_sheet(wb, K.SHEET_ITEM_LIST)
    inventory = find_sheet(wb, K.SHEET_INVENTORY)
    shopping = find_sheet(wb, K.SHEET_SHOPPING_HISTORY)

    indices = {
        K.SHEET_ITEM_LIST: _build_compile_index(item_list, "SKU CODE", loose=False),
        K.SHEET_INVENTORY: _build_compile_index(inventory, "Item", loose=False),
        K.SHEET_SHOPPING_HISTORY: _build_compile_index(shopping, "Item", loose=True),
    }

    # Write target headers at the spec-defined column letters.
    for _src, _src_col, tgt_letter, tgt_header in K.COMPILE_MAP:
        all_orders[f"{tgt_letter}1"] = tgt_header

    code_col = header_index(all_orders, "Code")
    # All Orders is the target, we assume it's clean but let's be safe
    for row_idx, values in iter_data_rows(all_orders, known_headers=["Code"]):
        raw_code = values[code_col - 1]
        key_strict = code_key(raw_code)
        key_loose = code_key_loose(raw_code)
        for src_sheet, src_header, tgt_letter, _ in K.COMPILE_MAP:
            idx = indices[src_sheet]
            if src_sheet == K.SHEET_SHOPPING_HISTORY:
                record = idx.get(key_loose)
            else:
                record = idx.get(key_strict)
            value = _record_get(record, src_header) if record else None
            all_orders[f"{tgt_letter}{row_idx}"] = value

    # Apply 0.#### number format to all precision columns.
    hmap = header_map(all_orders)
    for header in K.PRECISION_COLUMNS:
        if header in hmap:
            apply_number_format(all_orders, hmap[header], K.NUMBER_FORMAT_PRECISION)


def _copy_row_to(ws_dst: Worksheet, values: list[Any]) -> int:
    """Append `values` to `ws_dst` and return the new row index."""
    new_row = ws_dst.max_row + 1
    for ci, v in enumerate(values, start=1):
        ws_dst.cell(row=new_row, column=ci, value=v)
    return new_row


def _clone_headers(ws_src: Worksheet, ws_dst: Worksheet) -> None:
    for cell in ws_src[1]:
        ws_dst.cell(row=1, column=cell.column, value=cell.value)


def extract_jetro_and_po(wb: Workbook) -> None:
    """Spec §3.2 + §3.3 — numeric-bin -> Jetro source ; QOH=0 -> PO."""
    src = find_sheet(wb, K.SHEET_ALL_ORDERS)
    jetro = wb.create_sheet(K.SHEET_JETRO_SOURCE)
    po = wb.create_sheet(K.SHEET_PO)
    
    # Ensure we use the detected header row for headers
    h_idx = find_header_row(src, "Bin", "Quantity On Hand", "Code")
    for cell in src[h_idx]:
        jetro.cell(row=1, column=cell.column, value=cell.value)
        po.cell(row=1, column=cell.column, value=cell.value)

    bin_col = header_index(src, "Bin")
    qoh_col = header_index(src, "Quantity On Hand")

    keep_rows: list[list[Any]] = []
    for row_idx, values in iter_data_rows(src, known_headers=["Bin", "Quantity On Hand"]):
        bin_val = values[bin_col - 1]
        qoh_val = values[qoh_col - 1]
        if _is_numeric_bin(bin_val):
            _copy_row_to(jetro, list(values))
            continue
        if qoh_val == 0:
            _copy_row_to(po, list(values))
            continue
        keep_rows.append(list(values))

    # Clear original sheet and write back kept rows
    # Start clearing from the data row onwards
    src.delete_rows(h_idx + 1, src.max_row)
    for r in keep_rows:
        _copy_row_to(src, r)
