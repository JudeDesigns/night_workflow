"""Part 4 — Jetro Workflow (spec §6)."""
from __future__ import annotations

import datetime
from collections import defaultdict
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from . import constants as K
from .codes import code_key
from .sheet_utils import (
    find_sheet,
    header_map,
    iter_data_rows,
    find_header_row,
    int_if_whole,
)

from .pdf_utils import render_keep_together_pdf
from .report_styles import (
    apply_z_driver_shading,
    is_z_driver,
    set_column_widths,
    style_column_header_row,
    style_data_row,
    style_group_band,
)

try:
    from weasyprint import HTML, CSS
except ImportError:
    HTML = None
    CSS = None

def build_jetro_branch(wb: Workbook, job_dir: str) -> dict[str, dict[str, str]]:
    """Spec §6 — Sort Jetro, build Jetro page, Produce sheet, and PDF report."""
    
    # 1. Sort Jetro Source (Spec §6.3)
    _sort_jetro_source(wb)
    
    outputs = {}
    
    # 2. Jetro Report PDF (Spec §6.5) - Run first to get page numbers
    jetro_pdf, page_map, blocks = _build_jetro_pdf(wb, job_dir)
    outputs["jetroPdf"] = {"pdf": jetro_pdf}
    
    # 3. Jetro Excel Page & Produce Sheet (Spec §6.4, §6.7, §6.6)
    jetro_xlsx = _build_jetro_excel(wb, job_dir, blocks, page_map)
    outputs["jetroWorkbook"] = {"xlsx": jetro_xlsx}
    
    return outputs

def _get_driver_sequence(wb: Workbook) -> list[str]:
    # Driver Setup has only 2 columns (Order, Freezer group), so iter_data_rows'
    # 4-cell minimum filter would skip every row. Read it directly instead.
    try:
        ws = find_sheet(wb, K.SHEET_DRIVER_SETUP)
        seq: list[str] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            val = row[0] if row else None
            if val not in (None, ""):
                seq.append(str(val).strip())
        if seq:
            return seq
    except KeyError:
        pass
    return list(K.DEFAULT_JETRO_DRIVER_ORDER)

def _sort_jetro_source(wb: Workbook) -> None:
    try:
        ws = find_sheet(wb, K.SHEET_JETRO_SOURCE)
        h_idx = find_header_row(ws, "Driver", "Name", "Bin")
        hmap = header_map(ws, known_headers=["Driver", "Name", "Bin"])
        driver_col = hmap.get("Driver")
        cust_col = hmap.get("Name")
        bin_col = hmap.get("Bin")
        
        if not (driver_col and cust_col and bin_col): return
        
        seq = _get_driver_sequence(wb)
        driver_rank = {name: i for i, name in enumerate(seq)}
        
        headers = [c.value for c in ws[h_idx]]
        rows = []
        for _, values in iter_data_rows(ws, known_headers=["Driver", "Name", "Bin"]):
            rows.append(list(values))
            
        def sort_key(r):
            # Strip both sides — Driver Setup stores trimmed names but source
            # rows often carry trailing whitespace. Without .strip() the rank
            # lookup returns 999 and every row is sorted to the end.
            d = str(r[driver_col-1] or "").strip()
            c = str(r[cust_col-1] or "").lower()
            try:
                b = float(r[bin_col-1] or 0)
            except (ValueError, TypeError):
                b = 999
            return (driver_rank.get(d, 999), c, b)
            
        rows.sort(key=sort_key)
        
        ws.delete_rows(h_idx, ws.max_row)
        ws.append(headers)
        for r in rows:
            ws.append(r)
    except KeyError:
        pass

def _build_jetro_excel(wb: Workbook, job_dir: str, blocks: list[dict[str, Any]], page_map: list[tuple[int, int]]) -> str:
    out_wb = Workbook()
    ws_jetro = out_wb.active
    ws_jetro.title = "Jetro Page"
    
    today = datetime.date.today().strftime("%m/%d/%Y")
    
    # Relevant columns from SHEET_JETRO_SOURCE
    src = find_sheet(wb, K.SHEET_JETRO_SOURCE)
    known = ["Qty", "Product Name", "Code", "Bin", "Description", "Name", "Transaction Date", "Driver", "CATEGORY NAME", "Product", "New bin", "Quantity On Hand", "BIN(Internal)", "BIN (Internal)", "internal bin", "BIN"]
    hmap = header_map(src, known_headers=known)
    
    qty_idx = hmap.get("Qty")
    pname_idx = hmap.get("Product Name")
    code_idx = hmap.get("Code")
    bin_idx = hmap.get("Bin")
    # Internal Bin mapping
    internal_bin_idx = hmap.get("BIN(Internal)") or hmap.get("BIN (Internal)") or hmap.get("internal bin") or hmap.get("BIN") or bin_idx
    desc_idx = hmap.get("Description")
    cust_idx = hmap.get("Name")
    cat_idx = hmap.get("CATEGORY NAME")
    prod_idx = hmap.get("Product")
    new_bin_idx = hmap.get("New bin")
    qoh_idx = hmap.get("Quantity On Hand")
    driver_idx = hmap.get("Driver")
        
    # PRD §6.4: col A is the numeric Jetro Bin (source column F).
    headers = ["Bin", "New bin", "Qty", "Product", "QTY OH"]
    n_cols = len(headers)
    center_cols = {1, 2, 3, 5}
    set_column_widths(ws_jetro, [10, 12, 8, 60, 10])
    current_row = 1

    for block in blocks:
        cname = block["customer"]
        rows = block["rows"]
        title = block["title"]

        ws_jetro.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=n_cols)
        ws_jetro.cell(row=current_row, column=1, value=title)
        style_group_band(ws_jetro, current_row, n_cols)
        current_row += 1

        for i, h in enumerate(headers, start=1):
            ws_jetro.cell(row=current_row, column=i, value=h)
        style_column_header_row(ws_jetro, current_row, n_cols)
        current_row += 1

        for r in rows:
            # PRD §6.4: Jetro Page col A = source column F ("Bin") — the
            # numeric Jetro bin (e.g. "6C", "12"), not the internal warehouse
            # bin (e.g. "FREEZER-A11"). New bin is blanked when it matches Bin.
            bin_val = r[bin_idx-1] if bin_idx and bin_idx <= len(r) else ""
            new_bin_val = r[new_bin_idx-1] if new_bin_idx and new_bin_idx <= len(r) else ""
            if str(new_bin_val).strip().upper() == str(bin_val).strip().upper():
                new_bin_val = ""

            # Product merge: ProductName/Description/Product/Code
            pname = r[pname_idx-1] if pname_idx and pname_idx <= len(r) else ""
            desc = r[desc_idx-1] if desc_idx and desc_idx <= len(r) else ""
            prod = r[prod_idx-1] if prod_idx and prod_idx <= len(r) else ""
            code = r[code_idx-1] if code_idx and code_idx <= len(r) else ""
            p_merge = f"{pname}/{desc}/{prod}/{code}".replace("//", "/").replace("//", "/")

            row_data = [
                bin_val,
                new_bin_val,
                r[qty_idx-1] if qty_idx and qty_idx <= len(r) else "",
                p_merge,
                r[qoh_idx-1] if qoh_idx and qoh_idx <= len(r) else ""
            ]
            for i, v in enumerate(row_data, start=1):
                ws_jetro.cell(row=current_row, column=i, value=v)
            style_data_row(ws_jetro, current_row, n_cols, center_cols=center_cols)

            # Produce shading overlays the base data styling (PRD §6.4).
            is_produce = str(r[cat_idx-1]).lower() == "produce" if cat_idx and cat_idx <= len(r) else False
            driver_val = r[driver_idx-1] if driver_idx and driver_idx <= len(r) else ""
            if is_produce:
                fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                for i in range(1, n_cols + 1):
                    ws_jetro.cell(row=current_row, column=i).fill = fill
            elif is_z_driver(driver_val):
                # Z driver = future delivery / pickup. Distinct cool-grey shade.
                apply_z_driver_shading(ws_jetro, current_row, n_cols)

            current_row += 1
        current_row += 1

    # --- Jetro Produce Sheet (Spec §6.7) ---
    # Per PRD §6.7: col A is "Bin" (source column F), the numeric Jetro bin.
    ws_prod = out_wb.create_sheet("Jetro Produce Sheet")
    prod_headers = ["Bin", "New bin", "Qty", "Total qty", "Product Info", "Sell price", "Cost price", "JTR U COST", "JTR C cost", "Customer name", "Driver name"]
    prod_n_cols = len(prod_headers)
    prod_center_cols = {1, 2, 3, 4, 6, 7, 8, 9}
    set_column_widths(ws_prod, [10, 10, 8, 10, 50, 12, 12, 12, 12, 28, 16])

    # Direct fallback lookup from Shopping History for JTR U/C COST. The Part 1
    # compile already propagates these into "unit price"/"case price" on Jetro
    # source, but if a produce code didn't match during the join we re-attempt
    # the lookup here so the Produce sheet doesn't silently render blanks.
    sh_index: dict[str, tuple[Any, Any]] = {}
    try:
        sh = find_sheet(wb, K.SHEET_SHOPPING_HISTORY)
        sh_hmap = header_map(sh, known_headers=["Item", "Unit Price", "Case Price"])
        sh_item_col = sh_hmap.get("Item")
        sh_up_col = sh_hmap.get("Unit Price")
        sh_cp_col = sh_hmap.get("Case Price")
        if sh_item_col and (sh_up_col or sh_cp_col):
            for _, sh_vals in iter_data_rows(sh, known_headers=["Item"]):
                raw = sh_vals[sh_item_col-1] if sh_item_col <= len(sh_vals) else None
                key = code_key(raw)
                if not key or key in sh_index:
                    continue
                up = sh_vals[sh_up_col-1] if sh_up_col and sh_up_col <= len(sh_vals) else None
                cp = sh_vals[sh_cp_col-1] if sh_cp_col and sh_cp_col <= len(sh_vals) else None
                sh_index[key] = (up, cp)
    except KeyError:
        pass

    def _sh_lookup(code_val: Any) -> tuple[Any, Any]:
        """Return (unit_price, case_price) from Shopping History, or (None, None)."""
        k = code_key(code_val)
        if k in sh_index:
            return sh_index[k]
        # Loose match: strip a trailing U/C from either side (Spec §2).
        from .codes import code_key_loose
        kl = code_key_loose(code_val)
        for sk, val in sh_index.items():
            if code_key_loose(sk) == kl:
                return val
        return (None, None)

    # Filter produce rows
    produce_rows = []
    for _, values in iter_data_rows(src):
        if cat_idx and cat_idx <= len(values) and str(values[cat_idx-1]).lower() == "produce":
            produce_rows.append(values)

    # Sort by Bin (col F), then Product Name (PRD §6.7).
    _b_sort = (bin_idx - 1) if bin_idx else -1
    produce_rows.sort(key=lambda x: (str(x[_b_sort] or "") if _b_sort >= 0 and _b_sort < len(x) else "",
                                     str(x[pname_idx-1] or "").lower() if pname_idx and pname_idx <= len(x) else ""))

    # Write headers
    for i, h in enumerate(prod_headers, start=1):
        ws_prod.cell(row=1, column=i, value=h)
    style_column_header_row(ws_prod, 1, prod_n_cols)
        
    # Calculate totals for repeated product names
    totals = defaultdict(float)
    for r in produce_rows:
        if pname_idx and pname_idx <= len(r) and qty_idx and qty_idx <= len(r):
            totals[str(r[pname_idx-1]).lower()] += float(r[qty_idx-1] or 0)
        
    seen_pnames = set()
    for i, r in enumerate(produce_rows, start=2):
        pname = str(r[pname_idx-1]).lower() if pname_idx and pname_idx <= len(r) else ""
        tqty = ""
        if pname not in seen_pnames:
            tqty = int_if_whole(totals[pname])
            seen_pnames.add(pname)
            
        bin_val = r[bin_idx-1] if bin_idx and bin_idx <= len(r) else ""
        new_bin_val = r[new_bin_idx-1] if new_bin_idx and new_bin_idx <= len(r) else ""
        # PRD §6.7: blank New bin when it equals the Jetro Bin (column F).
        if str(new_bin_val).strip().upper() == str(bin_val).strip().upper(): new_bin_val = ""

        # Product Info: ProductName/Product/Code
        p_info = f"{r[pname_idx-1] if pname_idx and pname_idx <= len(r) else ''}/{r[prod_idx-1] if prod_idx and prod_idx <= len(r) else ''}/{r[code_idx-1] if code_idx and code_idx <= len(r) else ''}".replace("//", "/")

        # JTR U/C COST — propagated from Shopping History via Part 1 compile.
        # Fall back to a direct Shopping History lookup if the propagated value
        # is missing (covers codes that didn't match during the Part 1 join).
        up_col = hmap.get("unit price")
        cp_col = hmap.get("case price")
        up_val = r[up_col-1] if up_col and up_col <= len(r) else None
        cp_val = r[cp_col-1] if cp_col and cp_col <= len(r) else None
        if (up_val in (None, "")) or (cp_val in (None, "")):
            code_for_lookup = r[code_idx-1] if code_idx and code_idx <= len(r) else None
            sh_up, sh_cp = _sh_lookup(code_for_lookup)
            if up_val in (None, ""): up_val = sh_up if sh_up is not None else ""
            if cp_val in (None, ""): cp_val = sh_cp if sh_cp is not None else ""

        row_data = [
            bin_val,
            new_bin_val,
            r[qty_idx-1] if qty_idx and qty_idx <= len(r) else "",
            tqty,
            p_info,
            # Sell price = Price (column M) per PRD §6.7/§6.1.
            (r[hmap.get("Price")-1] if hmap.get("Price") and hmap.get("Price") <= len(r) else ""),
            r[hmap.get("CURRENT COST PRICE")-1] if hmap.get("CURRENT COST PRICE") and hmap.get("CURRENT COST PRICE") <= len(r) else (r[hmap.get("Cost")-1] if hmap.get("Cost") and hmap.get("Cost") <= len(r) else ""),
            up_val if up_val is not None else "",
            cp_val if cp_val is not None else "",
            r[cust_idx-1] if cust_idx and cust_idx <= len(r) else "",
            r[hmap.get("Driver")-1] if hmap.get("Driver") and hmap.get("Driver") <= len(r) else ""
        ]
        for ci, v in enumerate(row_data, start=1):
            cell = ws_prod.cell(row=i, column=ci, value=v)
            if ci in (6, 7, 8, 9): cell.number_format = K.NUMBER_FORMAT_PRECISION
        style_data_row(ws_prod, i, prod_n_cols, center_cols=prod_center_cols)
        # Driver column is the last element of row_data; flag Z-driver rows.
        if is_z_driver(row_data[-1]):
            apply_z_driver_shading(ws_prod, i, prod_n_cols)

    # --- Menu Page (Spec §6.6) ---
    # page_map holds (start_page, end_page) per block. Orders that span more
    # than one page show a range like "3-6"; single-page orders show "3".
    ws_menu = out_wb.create_sheet("Menu Page")
    menu_headers = ["Page number", "Customer Orders", "Pulled by"]
    menu_n_cols = len(menu_headers)
    set_column_widths(ws_menu, [16, 50, 28])
    for i, h in enumerate(menu_headers, start=1):
        ws_menu.cell(row=1, column=i, value=h)
    style_column_header_row(ws_menu, 1, menu_n_cols)
    for ridx, (block, pages) in enumerate(zip(blocks, page_map), start=2):
        start, end = pages
        page_label = f"{start}-{end}" if end > start else f"{start}"
        ws_menu.cell(row=ridx, column=1, value=page_label)
        ws_menu.cell(row=ridx, column=2, value=block["customer"])
        ws_menu.cell(row=ridx, column=3, value="")
        style_data_row(ws_menu, ridx, menu_n_cols, center_cols={1})
    
    xlsx_path = f"{job_dir}/jetro.xlsx"
    out_wb.save(xlsx_path)
    return "jetro.xlsx"

def _build_jetro_pdf(wb: Workbook, job_dir: str) -> tuple[str, list[tuple[int, int]], list[dict[str, Any]]]:
    if not HTML: return "", [], []
    
    src = find_sheet(wb, K.SHEET_JETRO_SOURCE)
    known = ["Qty", "Product Name", "Code", "Bin", "Description", "Name", "Transaction Date", "Driver", "CATEGORY NAME", "Product", "New bin", "Quantity On Hand", "BIN(Internal)", "BIN (Internal)", "internal bin", "BIN"]
    hmap = header_map(src, known_headers=known)
    
    qty_idx = hmap.get("Qty")
    pname_idx = hmap.get("Product Name")
    code_idx = hmap.get("Code")
    bin_idx = hmap.get("Bin")
    internal_bin_idx = hmap.get("BIN(Internal)") or hmap.get("BIN (Internal)") or hmap.get("internal bin") or hmap.get("BIN") or bin_idx
    desc_idx = hmap.get("Description")
    cust_idx = hmap.get("Name")
    date_idx = hmap.get("Transaction Date")
    driver_idx = hmap.get("Driver")
    cat_idx = hmap.get("CATEGORY NAME")
    prod_idx = hmap.get("Product")
    new_bin_idx = hmap.get("New bin")
    qoh_idx = hmap.get("Quantity On Hand")
    
    today = datetime.date.today().strftime("%m/%d/%Y")
    
    # Group by customer
    blocks = []
    current_cust = None
    cust_rows = []
    for _, values in iter_data_rows(src, known_headers=known):
        cname = str(values[cust_idx-1] or "Unknown") if cust_idx else "Unknown"
        if cname != current_cust:
            if current_cust:
                date_val = cust_rows[0][date_idx-1] if date_idx else today
                driver_val = cust_rows[0][driver_idx-1] if driver_idx else ""
                blocks.append({
                    "title": f"{current_cust} | {date_val} | {driver_val}",
                    "rows": cust_rows,
                    "customer": current_cust
                })
            current_cust = cname
            cust_rows = []
        cust_rows.append(values)
    if current_cust:
        date_val = cust_rows[0][date_idx-1] if date_idx else today
        driver_val = cust_rows[0][driver_idx-1] if driver_idx else ""
        blocks.append({
            "title": f"{current_cust} | {date_val} | {driver_val}",
            "rows": cust_rows,
            "customer": current_cust
        })

    def template(block, font_size):
        rows_html = ""
        for r in block["rows"]:
            # PRD §6.4: col A is the numeric Jetro Bin (source column F), matching
            # the Excel Jetro Page. New bin is blanked when it equals Bin.
            bin_val = r[bin_idx-1] if bin_idx and bin_idx <= len(r) else ""
            new_bin_val = r[new_bin_idx-1] if new_bin_idx and new_bin_idx <= len(r) else ""
            if str(new_bin_val).strip().upper() == str(bin_val).strip().upper(): new_bin_val = ""

            pname = r[pname_idx-1] if pname_idx and pname_idx <= len(r) else ""
            desc = r[desc_idx-1] if desc_idx and desc_idx <= len(r) else ""
            prod = r[prod_idx-1] if prod_idx and prod_idx <= len(r) else ""
            code = r[code_idx-1] if code_idx and code_idx <= len(r) else ""
            p_merge = f"{pname}/{desc}/{prod}/{code}".replace("//", "/")

            is_produce = str(r[cat_idx-1]).lower() == "produce" if cat_idx and cat_idx <= len(r) else False

            rows_html += f"""
                <tr class="{'produce' if is_produce else ''}">
                    <td>{bin_val}</td>
                    <td>{new_bin_val}</td>
                    <td class="centered">{r[qty_idx-1] if qty_idx and qty_idx <= len(r) else ""}</td>
                    <td>{p_merge}</td>
                    <td class="centered">{r[qoh_idx-1] if qoh_idx and qoh_idx <= len(r) else ""}</td>
                </tr>
            """

        return f"""
        <div style="font-size: {font_size}pt;">
            <style>
                .produce {{ background-color: #f9f9f9; }}
                .centered {{ text-align: center; }}
            </style>
            <table>
                <colgroup>
                    <col style="width:7%">
                    <col style="width:7%">
                    <col style="width:5%">
                    <col style="width:74%">
                    <col style="width:7%">
                </colgroup>
                <thead>
                    <tr><th colspan="5" class="header">{block['title']}</th></tr>
                    <tr>
                        <th class="centered">Bin</th>
                        <th class="centered">New bin</th>
                        <th class="centered">Qty</th>
                        <th>Product</th>
                        <th class="centered">QTY OH</th>
                    </tr>
                </thead>
                <tbody>
                    {rows_html}
                </tbody>
            </table>
        </div>
        """
        
    pdf_path = f"{job_dir}/jetro_report.pdf"
    page_map = render_keep_together_pdf(blocks, pdf_path, template, landscape=True)
    
    return "jetro_report.pdf", page_map, blocks
