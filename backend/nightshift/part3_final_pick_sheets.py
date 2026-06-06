"""Part 3 — Final Pick Sheets (spec §5)."""
from __future__ import annotations

import datetime
from collections import defaultdict
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment, Font, Border, Side

from . import constants as K
from .codes import code_key
from .sheet_utils import (
    find_sheet,
    header_map,
    iter_data_rows,
    find_header_row,
)

from .pdf_utils import render_keep_together_pdf
from .report_styles import (
    apply_z_driver_shading,
    is_z_driver,
    set_column_widths,
    style_column_header_row,
    style_data_row,
    style_group_band,
)

try:
    from weasyprint import HTML, CSS
except ImportError:
    HTML = None
    CSS = None

# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def build_final_pick_sheets(wb: Workbook, job_dir: str) -> dict[str, dict[str, str]]:
    """Spec §5 — Build Excel and PDF reports for PO, Dry, Freezer, and WH Pickup.

    Returns an outputs map where the Dry / Freezer / WH Pickup PDFs are
    emitted as separate files alongside the combined pick_sheets.xlsx, per
    operational request (so the warehouse can hand each driver only the
    sheets they need).
    """

    # 1. Re-order All Orders (Spec §5.1)
    _reorder_all_orders(wb)

    outputs = {}

    # 2. Block 1 — PO by Vendors (Spec §5.2)
    po_xlsx, po_pdf = _build_po_report(wb, job_dir)
    outputs["poReport"] = {"xlsx": po_xlsx, "pdf": po_pdf}

    # 3. Block 2, 3, 4 — Dry, Freezer, WH Pickup. One combined xlsx; three
    # separate PDFs so each can be printed and handed out independently.
    pdf_paths = _build_combined_pick_sheets(wb, job_dir)
    outputs["dryFreezerWh"] = {
        "xlsx": pdf_paths["xlsx"],
        "dryPdf": pdf_paths["dryPdf"],
        "freezerPdf": pdf_paths["freezerPdf"],
        "whPickupPdf": pdf_paths["whPickupPdf"],
    }

    return outputs

# ---------------------------------------------------------------------------
# Logic
# ---------------------------------------------------------------------------

def _get_driver_sequence(wb: Workbook) -> list[str]:
    # Driver Setup has only 2 columns (Order, Freezer group), so iter_data_rows'
    # 4-cell minimum filter would skip every row. Read it directly instead.
    try:
        ws = find_sheet(wb, K.SHEET_DRIVER_SETUP)
        seq: list[str] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            val = row[0] if row else None
            if val not in (None, ""):
                seq.append(str(val).strip())
        if seq:
            return seq
    except KeyError:
        pass
    return list(K.DEFAULT_JETRO_DRIVER_ORDER)

def _reorder_all_orders(wb: Workbook) -> None:
    """Sort All Orders so rows follow the driver pull sequence."""
    try:
        ws = find_sheet(wb, K.SHEET_ALL_ORDERS)
        h_idx = find_header_row(ws, "Driver")
        hmap = header_map(ws, known_headers=["Driver"])
        driver_col = hmap.get("Driver")
        if not driver_col: return
        
        seq = _get_driver_sequence(wb)
        driver_rank = {name: i for i, name in enumerate(seq)}
        
        headers = [c.value for c in ws[h_idx]]
        rows = []
        for _, values in iter_data_rows(ws, known_headers=["Driver"]):
            rows.append(list(values))
            
        def sort_key(r):
            d = str(r[driver_col - 1] or "")
            return (driver_rank.get(d, 999), d)
            
        rows.sort(key=sort_key)
        
        ws.delete_rows(h_idx, ws.max_row)
        ws.append(headers)
        for r in rows:
            ws.append(r)
    except KeyError:
        pass

# --- PO Report ---

def _build_po_report(wb: Workbook, job_dir: str) -> tuple[str, str]:
    po_ws = find_sheet(wb, K.SHEET_PO)
    h_idx = find_header_row(po_ws, "Vendor", "Code")
    hmap = header_map(po_ws, known_headers=["Vendor", "Code"])
    
    # Group by vendor
    vendor_groups = defaultdict(list)
    vendor_col = hmap.get("Vendor")
    code_col = hmap.get("Code")
    
    if not vendor_col or not code_col:
        # Fallback to defaults if headers missing
        vendor_col = vendor_col or 8
        code_col = code_col or 5
    
    for _, values in iter_data_rows(po_ws, known_headers=["Vendor", "Code"]):
        v_idx = vendor_col - 1
        v = str(values[v_idx] or "Unknown").replace(",", "").strip() if v_idx < len(values) else "Unknown"
        vendor_groups[v].append(values)
        
    # Sort within each vendor by item code
    for v in vendor_groups:
        c_idx = code_col - 1
        vendor_groups[v].sort(key=lambda x: str(x[c_idx] or "") if c_idx < len(x) else "")
        
    # Create Excel
    out_wb = Workbook()
    ws = out_wb.active
    ws.title = "PO by Vendors"
    
    today = datetime.date.today().strftime("%m/%d/%Y")
    
    headers = ["Qty", "total QTY", "Product Name", "Code", "Price", "CURRENT COST PRICE", "Jetro cost", "Name", "Driver"]
    # Mapping to original PO sheet columns
    # A: Sheet, B: Vendor (route), C: Total Qty, D...
    # PO Sheet layout:
    # 0: Sheet, 1: Vendor(route), 2: Total Qty, 3: Product Name, 4: Code, 5: Bin, 6: Description, 7: Vendor, 8: Address, 9: Qty, ...
    # From spec §5.2:
    # Col A: Qty (from PO.Qty)
    # Col B: total QTY (from PO.Total Qty)
    # Col C: Product Name
    # Col D: Code
    # Col E: Price (Selling Price)
    # Col F: CURRENT COST PRICE
    # Col G: Jetro cost (Case Price)
    # Col H: Name (Customer)
    # Col I: Driver
    
    # Let's find indices in the original PO sheet
    qty_idx = hmap.get("Qty")
    total_qty_idx = 3 # Column C
    pname_idx = hmap.get("Product Name")
    code_idx = hmap.get("Code")
    price_idx = hmap.get("Price")
    cost_idx = hmap.get("CURRENT COST PRICE") or hmap.get("Cost") # Try 'Cost' as fallback
    jcost_idx = hmap.get("case price")
    name_idx = hmap.get("Name")
    driver_idx = hmap.get("Driver")

    n_cols = len(headers)
    # Qty / Total Qty / Price / Cost / Jetro cost — all centered per spec §10.
    center_cols = {1, 2, 5, 6, 7}
    set_column_widths(ws, [8, 10, 50, 12, 12, 14, 12, 28, 14])

    current_row = 1
    formatted_vendor_groups = defaultdict(list)

    for vendor in sorted(vendor_groups.keys()):
        # Spanned header
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=n_cols)
        ws.cell(row=current_row, column=1, value=f"{vendor}   \u2014   {today}")
        style_group_band(ws, current_row, n_cols)
        current_row += 1

        # Column headers
        for i, h in enumerate(headers, start=1):
            ws.cell(row=current_row, column=i, value=h)
        style_column_header_row(ws, current_row, n_cols)
        current_row += 1

        # Rows
        for vals in vendor_groups[vendor]:
            row_data = [
                vals[qty_idx - 1] if qty_idx else "",
                vals[total_qty_idx - 1] if len(vals) >= total_qty_idx else "",
                vals[pname_idx - 1] if pname_idx else "",
                vals[code_idx - 1] if code_idx else "",
                vals[price_idx - 1] if price_idx else "",
                vals[cost_idx - 1] if cost_idx else "",
                vals[jcost_idx - 1] if jcost_idx else "",
                vals[name_idx - 1] if name_idx else "",
                vals[driver_idx - 1] if driver_idx else "",
            ]
            is_z = is_z_driver(row_data[8])
            formatted_vendor_groups[vendor].append({"cells": row_data, "z": is_z})
            for i, v in enumerate(row_data, start=1):
                ws.cell(row=current_row, column=i, value=v)
            style_data_row(ws, current_row, n_cols, center_cols=center_cols)
            if is_z:
                apply_z_driver_shading(ws, current_row, n_cols)
            current_row += 1
        current_row += 1 # Gap between vendors

    xlsx_path = f"{job_dir}/po_report.xlsx"
    out_wb.save(xlsx_path)
    
    # PDF generation — center Qty/Price columns; pin column widths so wide
    # product names cannot overflow the page (spec §10 keep-together).
    pdf_path = f"{job_dir}/po_report.pdf"
    po_widths = ["6%", "7%", "30%", "10%", "8%", "11%", "8%", "14%", "6%"]
    _generate_pdf_generic(
        formatted_vendor_groups, headers, "PO Report", pdf_path, today,
        center_cols=center_cols, col_widths_pct=po_widths,
    )

    return "po_report.xlsx", "po_report.pdf"

def _build_combined_pick_sheets(wb: Workbook, job_dir: str) -> dict[str, str]:
    """One xlsx (3 sheets: Dry / Freezer / WH Pickup) plus three separate PDFs
    so each can be printed and handed out independently."""
    out_wb = Workbook()

    today = datetime.date.today().strftime("%m/%d/%Y")

    _add_dry_sheet(wb, out_wb, today)
    _add_freezer_sheet(wb, out_wb, today)
    _add_wh_pickup_sheet(wb, out_wb, today)

    # Workbook() seeds a blank "Sheet" tab; drop it now that the real
    # sheets are in place so the file opens on Dry by Driver.
    default = out_wb["Sheet"] if "Sheet" in out_wb.sheetnames else None
    if default is not None and len(out_wb.sheetnames) > 1:
        out_wb.remove(default)

    xlsx_path = f"{job_dir}/pick_sheets.xlsx"
    out_wb.save(xlsx_path)

    # Separate PDFs (operational request: one file per pick area).
    _generate_dry_pdf(wb, f"{job_dir}/dry.pdf", today)
    _generate_freezer_pdf(wb, f"{job_dir}/freezer.pdf", today)
    _generate_wh_pickup_pdf(wb, f"{job_dir}/wh_pickup.pdf", today)

    return {
        "xlsx": "pick_sheets.xlsx",
        "dryPdf": "dry.pdf",
        "freezerPdf": "freezer.pdf",
        "whPickupPdf": "wh_pickup.pdf",
    }

def _add_dry_sheet(wb: Workbook, out_wb: Workbook, today: str) -> None:
    ws_src = find_sheet(wb, K.SHEET_ALL_ORDERS)
    hmap = header_map(ws_src)
    bin_col = hmap.get("Bin")
    driver_col = hmap.get("Driver")
    internal_bin_col = hmap.get("BIN(Internal)") or hmap.get("BIN (Internal)") or hmap.get("internal bin") or hmap.get("BIN")
    
    if not (bin_col and driver_col): return
    
    # Filter for DRY
    dry_rows = []
    for _, values in iter_data_rows(ws_src):
        bin_val = str(values[bin_col - 1] or "").strip().upper()
        if bin_val == "DRY":
            dry_rows.append(values)
            
    # Group by driver (normalize names: source rows often carry trailing
    # whitespace that the Driver Setup sequence does not).
    driver_seq = [str(d).strip() for d in _get_driver_sequence(wb)]
    groups = defaultdict(list)
    for r in dry_rows:
        groups[str(r[driver_col - 1] or "Unknown").strip()].append(r)
    # Append any drivers present on rows but missing from the setup sequence.
    for extra in groups:
        if extra not in driver_seq:
            driver_seq.append(extra)
        
    ws = out_wb.create_sheet("Dry by Driver")
    # PRD §5.3 — Bin column dropped (always "DRY" inside this report).
    headers = ["Internal bin", "QTY", "Product Name", "Vendor", "Customer", "Driver", "QTY OH"]
    n_cols = len(headers)
    center_cols = {2, 7}
    set_column_widths(ws, [12, 8, 50, 28, 28, 14, 10])

    pname_idx = hmap.get("Product Name")
    desc_idx = hmap.get("Description")
    qty_idx = hmap.get("Qty")
    vendor_idx = hmap.get("Vendor")
    name_idx = hmap.get("Name")
    qoh_idx = hmap.get("Quantity On Hand")

    current_row = 1
    for driver in driver_seq:
        if driver not in groups: continue

        # Sort by internal bin
        rows = groups[driver]
        ib_idx = (internal_bin_col - 1) if internal_bin_col else -1
        rows.sort(key=lambda x: str(x[ib_idx] or "") if ib_idx >= 0 and ib_idx < len(x) else "")

        # Header
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=n_cols)
        ws.cell(row=current_row, column=1, value=f"{driver}   \u2014   {today}")
        style_group_band(ws, current_row, n_cols)
        current_row += 1

        for i, h in enumerate(headers, start=1):
            ws.cell(row=current_row, column=i, value=h)
        style_column_header_row(ws, current_row, n_cols)
        current_row += 1

        driver_is_z = is_z_driver(driver)
        for r in rows:
            pn_val = str(r[pname_idx-1] or "") if pname_idx and pname_idx <= len(r) else ""
            ds_val = str(r[desc_idx-1] or "") if desc_idx and desc_idx <= len(r) else ""
            pname = (pn_val + " " + ds_val).strip()

            row_data = [
                r[internal_bin_col-1] if internal_bin_col and internal_bin_col <= len(r) else "",
                r[qty_idx-1] if qty_idx and qty_idx <= len(r) else "",
                pname,
                r[vendor_idx-1] if vendor_idx and vendor_idx <= len(r) else "",
                r[name_idx-1] if name_idx and name_idx <= len(r) else "",
                r[driver_col-1] if driver_col and driver_col <= len(r) else "",
                r[qoh_idx-1] if qoh_idx and qoh_idx <= len(r) else "",
            ]
            for i, v in enumerate(row_data, start=1):
                ws.cell(row=current_row, column=i, value=v)
            style_data_row(ws, current_row, n_cols, center_cols=center_cols)
            if driver_is_z:
                apply_z_driver_shading(ws, current_row, n_cols)
            current_row += 1
        current_row += 1

def _add_freezer_sheet(wb: Workbook, out_wb: Workbook, today: str) -> None:
    ws_src = find_sheet(wb, K.SHEET_ALL_ORDERS)
    hmap = header_map(ws_src)
    bin_col = hmap.get("Bin")
    driver_col = hmap.get("Driver")
    pname_col = hmap.get("Product Name")
    code_col = hmap.get("Code")
    qty_col = hmap.get("Qty")
    internal_bin_col = hmap.get("BIN(Internal)") or hmap.get("BIN (Internal)") or hmap.get("internal bin") or hmap.get("BIN")
    name_col = hmap.get("Name")
    qoh_col = hmap.get("Quantity On Hand")
    
    if not bin_col: return
    
    freezer_rows = []
    for _, values in iter_data_rows(ws_src):
        bin_val = str(values[bin_col - 1] or "").strip().upper()
        if bin_val == "FREEZER":
            freezer_rows.append(values)
            
    driver_seq = [str(d).strip() for d in _get_driver_sequence(wb)]
    f1_drivers = set(driver_seq[:3])
    
    f1_rows = [r for r in freezer_rows if str(r[driver_col-1] if driver_col else "").strip() in f1_drivers]
    f2_rows = [r for r in freezer_rows if str(r[driver_col-1] if driver_col else "").strip() not in f1_drivers]
    
    ws = out_wb.create_sheet("Freezer Page")
    headers = ["Internal bin", "QTY", "Total Qty", "Product Name", "Customer", "QTY OH"]
    n_cols = len(headers)
    center_cols = {2, 3, 6}
    set_column_widths(ws, [14, 8, 10, 50, 28, 10])

    current_row = 1
    for group_name, rows in [("Freezer one", f1_rows), ("Freezer two", f2_rows)]:
        if not rows: continue

        # Sort by Product Name
        pn_idx = (pname_col - 1) if pname_col else -1
        rows.sort(key=lambda x: str(x[pn_idx] or "").lower() if pn_idx >= 0 and pn_idx < len(x) else "")

        # Header
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=n_cols)
        ws.cell(row=current_row, column=1, value=f"{group_name}   \u2014   {today}")
        style_group_band(ws, current_row, n_cols)
        current_row += 1

        for i, h in enumerate(headers, start=1):
            ws.cell(row=current_row, column=i, value=h)
        style_column_header_row(ws, current_row, n_cols)
        current_row += 1

        # Calculate totals per code within group
        totals = defaultdict(float)
        c_idx = (code_col - 1) if code_col else -1
        q_idx = (qty_col - 1) if qty_col else -1
        for r in rows:
            code_val = r[c_idx] if c_idx >= 0 and c_idx < len(r) else None
            key = code_key(code_val)
            qty_val = r[q_idx] if q_idx >= 0 and q_idx < len(r) else 0
            totals[key] += float(qty_val or 0)

        seen_codes = set()
        for r in rows:
            code_val = r[c_idx] if c_idx >= 0 and c_idx < len(r) else None
            key = code_key(code_val)
            tqty = totals[key] if key not in seen_codes else ""
            seen_codes.add(key)

            row_data = [
                r[internal_bin_col-1] if internal_bin_col and internal_bin_col <= len(r) else "",
                r[qty_col-1] if qty_col and qty_col <= len(r) else "",
                tqty,
                r[pname_col-1] if pname_col and pname_col <= len(r) else "",
                r[name_col-1] if name_col and name_col <= len(r) else "",
                r[qoh_col-1] if qoh_col and qoh_col <= len(r) else "",
            ]
            for i, v in enumerate(row_data, start=1):
                ws.cell(row=current_row, column=i, value=v)
            style_data_row(ws, current_row, n_cols, center_cols=center_cols)
            driver_val = r[driver_col-1] if driver_col and driver_col <= len(r) else ""
            if is_z_driver(driver_val):
                apply_z_driver_shading(ws, current_row, n_cols)
            current_row += 1
        current_row += 1

def _add_wh_pickup_sheet(wb: Workbook, out_wb: Workbook, today: str) -> None:
    ws_src = find_sheet(wb, K.SHEET_ALL_ORDERS)
    hmap = header_map(ws_src)
    name_col = hmap.get("Name")
    driver_col = hmap.get("Driver")
    bin_col = hmap.get("Bin")
    internal_bin_col = hmap.get("BIN(Internal)") or hmap.get("BIN (Internal)") or hmap.get("internal bin") or hmap.get("BIN")
    qty_col = hmap.get("Qty")
    pname_col = hmap.get("Product Name")
    desc_col = hmap.get("Description")
    vendor_col = hmap.get("Vendor")
    qoh_col = hmap.get("Quantity On Hand")
    
    if not name_col: return
    
    # Group by customer
    customers = defaultdict(list)
    for _, values in iter_data_rows(ws_src):
        customers[str(values[name_col-1] or "Unknown")].append(values)
        
    driver_seq = _get_driver_sequence(wb)
    driver_rank = {name: i for i, name in enumerate(driver_seq)}
    
    # Sort customers by driver pull sequence
    sorted_customers = sorted(customers.keys(), key=lambda c: driver_rank.get(str(customers[c][0][driver_col-1] if driver_col else ""), 999))
    
    ws = out_wb.create_sheet("WH Pickup")
    headers = ["QTY", "Product Name", "Bin", "Internal bin", "Vendor", "Driver", "QTY OH"]
    n_cols = len(headers)
    center_cols = {1, 7}
    set_column_widths(ws, [8, 50, 10, 14, 28, 14, 10])

    current_row = 1
    for cust in sorted_customers:
        rows = customers[cust]
        driver = str(rows[0][driver_col-1] or "") if driver_col and driver_col <= len(rows[0]) else ""

        # Sort rows: Cooler -> Cooler-pd -> Dry -> Freezer -> others -> then by Internal bin
        def row_sort_key(r):
            b_val = str(r[bin_col-1] or "").strip().upper() if bin_col and bin_col <= len(r) else ""
            priority = K.WH_PICKUP_BIN_ORDER.get(b_val, 99)
            ib_val = str(r[internal_bin_col-1] or "") if internal_bin_col and internal_bin_col <= len(r) else ""
            return (priority, b_val, ib_val)

        rows.sort(key=row_sort_key)

        # Header
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=n_cols)
        ws.cell(row=current_row, column=1, value=f"{cust}   \u2014   {driver}   \u2014   {today}")
        style_group_band(ws, current_row, n_cols)
        current_row += 1

        for i, h in enumerate(headers, start=1):
            ws.cell(row=current_row, column=i, value=h)
        style_column_header_row(ws, current_row, n_cols)
        current_row += 1

        for r in rows:
            pn_val = str(r[pname_col-1] or "") if pname_col and pname_col <= len(r) else ""
            ds_val = str(r[desc_col-1] or "") if desc_col and desc_col <= len(r) else ""
            pname = (pn_val + " " + ds_val).strip()

            driver_val = r[driver_col-1] if driver_col and driver_col <= len(r) else ""
            row_data = [
                r[qty_col-1] if qty_col and qty_col <= len(r) else "",
                pname,
                r[bin_col-1] if bin_col and bin_col <= len(r) else "",
                r[internal_bin_col-1] if internal_bin_col and internal_bin_col <= len(r) else "",
                r[vendor_col-1] if vendor_col and vendor_col <= len(r) else "",
                driver_val,
                r[qoh_col-1] if qoh_col and qoh_col <= len(r) else "",
            ]
            for i, v in enumerate(row_data, start=1):
                ws.cell(row=current_row, column=i, value=v)
            style_data_row(ws, current_row, n_cols, center_cols=center_cols)
            if is_z_driver(driver_val):
                apply_z_driver_shading(ws, current_row, n_cols)
            current_row += 1
        current_row += 1

# ---------------------------------------------------------------------------
# PDF Generation Helpers
# ---------------------------------------------------------------------------

def _generate_pdf_generic(groups, headers, title, output_path, date, center_cols=None, col_widths_pct=None):
    """Generic PDF generator. center_cols is a 1-based set of column indices to
    horizontally center. col_widths_pct is a list of % strings matching headers
    length (e.g. ['10%', '20%', ...])."""
    if not HTML: return

    center_cols = set(center_cols or [])

    blocks = []
    for group_name in sorted(groups.keys()):
        blocks.append({
            "title": f"{group_name} - {date}",
            "headers": headers,
            "rows": groups[group_name],
            "center_cols": center_cols,
            "col_widths_pct": col_widths_pct,
        })

    render_keep_together_pdf(blocks, output_path, _block_template)


def _block_template(block, font_size):
    """Shared HTML template for both _generate_pdf_generic and the combined
    PDF. Supports per-column centering + widths and `tr.z-driver` shading on
    rows where the office's selected driver is "Z" (future delivery)."""
    hdrs = block["headers"]
    c_cols = block.get("center_cols") or set()
    widths = block.get("col_widths_pct") or []
    col_html = ""
    if widths and len(widths) == len(hdrs):
        col_html = "<colgroup>" + "".join(f'<col style="width:{w}">' for w in widths) + "</colgroup>"
    header_cells = " ".join(
        f'<th class="centered">{h}</th>' if (i + 1) in c_cols else f"<th>{h}</th>"
        for i, h in enumerate(hdrs)
    )
    rows_html = ""
    for r in block["rows"]:
        cell_vals = r["cells"] if isinstance(r, dict) else r
        is_z = bool(isinstance(r, dict) and r.get("z"))
        tr_class = ' class="z-driver"' if is_z else ""
        cells = []
        for i, v in enumerate(cell_vals[:len(hdrs)]):
            txt = "" if v is None else v
            if (i + 1) in c_cols:
                cells.append(f'<td class="centered">{txt}</td>')
            else:
                cells.append(f"<td>{txt}</td>")
        rows_html += f"<tr{tr_class}>" + "".join(cells) + "</tr>"

    return f"""
    <div style="font-size: {font_size}pt;">
        <table>
            {col_html}
            <thead>
                <tr><th colspan="{len(hdrs)}" class="header">{block['title']}</th></tr>
                <tr>{header_cells}</tr>
            </thead>
            <tbody>
                {rows_html}
            </tbody>
        </table>
    </div>
    """


def _generate_dry_pdf(wb, output_path, date):
    """PRD §5.3 — Dry by Driver. One driver group per block. Bin column dropped
    (every row is DRY by definition); QTY + QTY OH centered."""
    if not HTML: return
    dry_groups = _get_dry_groups(wb)
    headers = ["Internal bin", "QTY", "Product Name", "Vendor", "Customer", "Driver", "QTY OH"]
    widths = ["11%", "6%", "33%", "18%", "18%", "10%", "4%"]
    center = {2, 7}
    blocks = [
        {"title": f"Dry: {driver} - {date}", "headers": headers, "rows": rows,
         "center_cols": center, "col_widths_pct": widths}
        for driver, rows in dry_groups.items()
    ]
    render_keep_together_pdf(blocks, output_path, _block_template)


def _generate_freezer_pdf(wb, output_path, date):
    """PRD §5.4 — Freezer Page. Freezer one / Freezer two are the atomic units."""
    if not HTML: return
    freezer_groups = _get_freezer_groups(wb)
    headers = ["Internal Bin", "qty", "Total Qty", "product name", "Customer", "Qty on Hand"]
    widths = ["13%", "7%", "9%", "44%", "20%", "7%"]
    center = {2, 3, 6}
    blocks = [
        {"title": f"Freezer: {group} - {date}", "headers": headers, "rows": rows,
         "center_cols": center, "col_widths_pct": widths}
        for group, rows in freezer_groups.items()
    ]
    render_keep_together_pdf(blocks, output_path, _block_template)


def _generate_wh_pickup_pdf(wb, output_path, date):
    """PRD §5.5 — WH Pickup. One customer block per page."""
    if not HTML: return
    wh_groups = _get_wh_pickup_groups(wb)
    headers = ["qty", "Product Name", "bin", "internal bin", "vendor", "Driver", "qty OH"]
    widths = ["6%", "36%", "10%", "12%", "18%", "12%", "6%"]
    center = {1, 7}
    blocks = [
        {"title": f"WH Pickup: {cust} - {date}", "headers": headers, "rows": rows,
         "center_cols": center, "col_widths_pct": widths}
        for cust, rows in wh_groups.items()
    ]
    render_keep_together_pdf(blocks, output_path, _block_template)

def _render_table(title, headers, rows):
    html = f"""
    <table>
        <thead>
            <tr><th colspan="{len(headers)}" class="header">{title}</th></tr>
            <tr>{" ".join(f"<th>{h}</th>" for h in headers)}</tr>
        </thead>
        <tbody>
    """
    for r in rows:
        html += "<tr>" + "".join(f"<td>{v if v is not None else ''}</td>" for v in r) + "</tr>"
    html += "</tbody></table>"
    return html

def _get_dry_groups(wb):
    """Returns {driver: [{"cells": [...], "z": bool}, ...]} so the PDF
    template can shade Z-driver rows distinctly."""
    ws = find_sheet(wb, K.SHEET_ALL_ORDERS)
    hmap = header_map(ws)
    bin_col = hmap.get("Bin")
    driver_col = hmap.get("Driver")
    internal_bin_col = hmap.get("BIN(Internal)") or hmap.get("BIN (Internal)") or hmap.get("internal bin") or hmap.get("BIN")
    pname_idx = hmap.get("Product Name")
    desc_idx = hmap.get("Description")
    qty_idx = hmap.get("Qty")
    vendor_idx = hmap.get("Vendor")
    name_idx = hmap.get("Name")
    qoh_idx = hmap.get("Quantity On Hand")

    groups = defaultdict(list)

    for _, r in iter_data_rows(ws):
        bin_val = str(r[bin_col-1] or "") if bin_col and bin_col <= len(r) else ""
        if bin_val.upper() == "DRY":
            driver = str(r[driver_col-1] or "") if driver_col and driver_col <= len(r) else "Unknown"

            pn_val = str(r[pname_idx-1] or "") if pname_idx and pname_idx <= len(r) else ""
            ds_val = str(r[desc_idx-1] or "") if desc_idx and desc_idx <= len(r) else ""
            pname = pn_val + " " + ds_val

            ib_val = r[internal_bin_col-1] if internal_bin_col and internal_bin_col <= len(r) else ""
            qty_val = r[qty_idx-1] if qty_idx and qty_idx <= len(r) else ""
            vendor_val = r[vendor_idx-1] if vendor_idx and vendor_idx <= len(r) else ""
            name_val = r[name_idx-1] if name_idx and name_idx <= len(r) else ""
            qoh_val = r[qoh_idx-1] if qoh_idx and qoh_idx <= len(r) else ""

            groups[driver].append({
                # PRD §5.3 — Bin column dropped (always "DRY" inside this report).
                "cells": [ib_val, qty_val, pname, vendor_val, name_val, driver, qoh_val],
                "z": is_z_driver(driver),
            })

    for d in groups:
        groups[d].sort(key=lambda x: str(x["cells"][0] or ""))
    return groups

def _get_freezer_groups(wb):
    ws = find_sheet(wb, K.SHEET_ALL_ORDERS)
    hmap = header_map(ws)
    bin_col = hmap.get("Bin")
    driver_col = hmap.get("Driver")
    pname_col = hmap.get("Product Name")
    code_col = hmap.get("Code")
    qty_col = hmap["Qty"]
    internal_bin_col = hmap.get("BIN(Internal)") or hmap.get("BIN (Internal)") or hmap.get("internal bin") or hmap.get("BIN")
    name_col = hmap.get("Name")
    qoh_col = hmap.get("Quantity On Hand")
    
    driver_seq = _get_driver_sequence(wb)
    f1_drivers = set(driver_seq[:3])
    
    groups = {"Freezer one": [], "Freezer two": []}
    for _, r in iter_data_rows(ws):
        bin_val = str(r[bin_col-1] or "") if bin_col and bin_col <= len(r) else ""
        if bin_val.upper() == "FREEZER":
            driver = str(r[driver_col-1]) if driver_col and driver_col <= len(r) else ""
            g = "Freezer one" if driver in f1_drivers else "Freezer two"
            groups[g].append(r)
            
    final_groups = {}
    for g, rows in groups.items():
        if not rows: continue
        
        # Sort by Product Name
        pn_idx = pname_col - 1 if pname_col else -1
        rows.sort(key=lambda x: str(x[pn_idx] or "").lower() if pn_idx >= 0 and pn_idx < len(x) else "")
        
        totals = defaultdict(float)
        c_idx = code_col - 1 if code_col else -1
        q_idx = qty_col - 1 if qty_col else -1
        for r in rows:
            c_val = r[c_idx] if c_idx >= 0 and c_idx < len(r) else None
            q_val = r[q_idx] if q_idx >= 0 and q_idx < len(r) else 0
            totals[code_key(c_val)] += float(q_val or 0)
            
        seen = set()
        formatted_rows = []
        for r in rows:
            c_val = r[c_idx] if c_idx >= 0 and c_idx < len(r) else None
            key = code_key(c_val)
            tqty = totals[key] if key not in seen else ""
            seen.add(key)
            d_val = str(r[driver_col-1]) if driver_col and driver_col <= len(r) else ""

            formatted_rows.append({
                "cells": [
                    r[internal_bin_col-1] if internal_bin_col and internal_bin_col <= len(r) else "",
                    r[qty_col-1] if qty_col and qty_col <= len(r) else "",
                    tqty,
                    r[pname_col-1] if pname_col and pname_col <= len(r) else "",
                    r[name_col-1] if name_col and name_col <= len(r) else "",
                    r[qoh_col-1] if qoh_col and qoh_col <= len(r) else "",
                ],
                "z": is_z_driver(d_val),
            })
        final_groups[g] = formatted_rows
    return final_groups

def _get_wh_pickup_groups(wb):
    ws = find_sheet(wb, K.SHEET_ALL_ORDERS)
    hmap = header_map(ws)
    name_col = hmap.get("Name")
    driver_col = hmap.get("Driver")
    bin_col = hmap.get("Bin")
    internal_bin_col = hmap.get("BIN(Internal)") or hmap.get("BIN (Internal)") or hmap.get("internal bin")
    qty_col = hmap.get("Qty")
    pname_col = hmap.get("Product Name")
    desc_col = hmap.get("Description")
    vendor_col = hmap.get("Vendor")
    qoh_col = hmap.get("Quantity On Hand")
    internal_bin_col = hmap.get("BIN(Internal)") or hmap.get("BIN (Internal)") or hmap.get("internal bin") or hmap.get("BIN")
    
    customers = defaultdict(list)
    for _, r in iter_data_rows(ws):
        cust_name = str(r[name_col-1] or "Unknown") if name_col and name_col <= len(r) else "Unknown"
        customers[cust_name].append(r)
        
    driver_seq = _get_driver_sequence(wb)
    driver_rank = {name: i for i, name in enumerate(driver_seq)}
    
    def cust_sort_key(c):
        rows = customers[c]
        d_val = str(rows[0][driver_col-1]) if driver_col and driver_col <= len(rows[0]) else ""
        return driver_rank.get(d_val, 999)
        
    sorted_custs = sorted(customers.keys(), key=cust_sort_key)
    
    final_groups = {}
    for cust in sorted_custs:
        rows = customers[cust]
        def row_sort_key(r):
            b_val = str(r[bin_col-1]).upper() if bin_col and bin_col <= len(r) else ""
            ib_val = str(r[internal_bin_col-1] or "") if internal_bin_col and internal_bin_col <= len(r) else ""
            return (K.WH_PICKUP_BIN_ORDER.get(b_val, 99), b_val, ib_val)
        rows.sort(key=row_sort_key)
        
        final_groups[cust] = []
        for r in rows:
            pn_val = str(r[pname_col-1] or "") if pname_col and pname_col <= len(r) else ""
            ds_val = str(r[desc_col-1] or "") if desc_col and desc_col <= len(r) else ""
            pname = pn_val + " " + ds_val
            d_val = r[driver_col-1] if driver_col and driver_col <= len(r) else ""

            final_groups[cust].append({
                "cells": [
                    r[qty_col-1] if qty_col and qty_col <= len(r) else "",
                    pname,
                    r[bin_col-1] if bin_col and bin_col <= len(r) else "",
                    r[internal_bin_col-1] if internal_bin_col and internal_bin_col <= len(r) else "",
                    r[vendor_col-1] if vendor_col and vendor_col <= len(r) else "",
                    d_val,
                    r[qoh_col-1] if qoh_col and qoh_col <= len(r) else "",
                ],
                "z": is_z_driver(d_val),
            })
    return final_groups
