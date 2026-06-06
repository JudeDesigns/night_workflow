"""Shared visual theme for the generated Excel report workbooks.

Used by part3 (PO / Dry / Freezer / WH Pickup) and part4 (Jetro page / Produce
/ Menu) so every output carries the same look: dark group bands, blue column
headers, bordered data rows.
"""
from __future__ import annotations

from typing import Any

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# Palette
COLOR_GROUP_BAND_BG = "1F2937"   # dark slate
COLOR_GROUP_BAND_FG = "FFFFFF"
COLOR_COL_HEADER_BG = "4472C4"   # office blue
COLOR_COL_HEADER_FG = "FFFFFF"
COLOR_BORDER = "BFBFBF"
# Cool blue-grey used to flag rows whose Driver = "Z" (future deliveries /
# pickups). Distinct from the produce light-grey (D3D3D3) used on Jetro pages.
COLOR_Z_DRIVER_BG = "B5BFC9"

_thin = Side(style="thin", color=COLOR_BORDER)
_DATA_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

_GROUP_BAND_FILL = PatternFill("solid", fgColor=COLOR_GROUP_BAND_BG)
_GROUP_BAND_FONT = Font(bold=True, color=COLOR_GROUP_BAND_FG, size=12)
# Spec §10: spanned headers (vendor/driver/customer + date) are horizontally centered.
_GROUP_BAND_ALIGN = Alignment(horizontal="center", vertical="center")

_COL_HEADER_FILL = PatternFill("solid", fgColor=COLOR_COL_HEADER_BG)
_COL_HEADER_FONT = Font(bold=True, color=COLOR_COL_HEADER_FG)
_COL_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")

_DATA_FONT = Font(color="000000")
_ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)
_ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
_ALIGN_CENTER = Alignment(horizontal="center", vertical="center")


def style_group_band(ws: Worksheet, row: int, n_cols: int) -> None:
    """Style a merged group-header band that already spans cols 1..n_cols."""
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = _GROUP_BAND_FILL
        cell.font = _GROUP_BAND_FONT
        cell.alignment = _GROUP_BAND_ALIGN
    ws.row_dimensions[row].height = 22


def style_column_header_row(ws: Worksheet, row: int, n_cols: int) -> None:
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = _COL_HEADER_FILL
        cell.font = _COL_HEADER_FONT
        cell.alignment = _COL_HEADER_ALIGN
        cell.border = _DATA_BORDER
    ws.row_dimensions[row].height = 20


def style_data_row(
    ws: Worksheet,
    row: int,
    n_cols: int,
    numeric_cols: set[int] | None = None,
    center_cols: set[int] | None = None,
) -> None:
    """Apply border + font + per-column alignment to a single data row."""
    numeric_cols = numeric_cols or set()
    center_cols = center_cols or set()
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = _DATA_FONT
        cell.border = _DATA_BORDER
        if c in center_cols:
            cell.alignment = _ALIGN_CENTER
        elif c in numeric_cols:
            cell.alignment = _ALIGN_RIGHT
        else:
            cell.alignment = _ALIGN_LEFT


def set_column_widths(ws: Worksheet, widths: list[int]) -> None:
    """Apply a list of column widths (1-indexed; 0 / None to skip)."""
    for i, w in enumerate(widths, start=1):
        if w:
            ws.column_dimensions[get_column_letter(i)].width = w


_Z_DRIVER_FILL = PatternFill("solid", fgColor=COLOR_Z_DRIVER_BG)


def apply_z_driver_shading(ws: Worksheet, row: int, n_cols: int) -> None:
    """Overlay the Z-driver grey fill on a data row that's already been
    `style_data_row`-styled. Cell borders and alignment are preserved."""
    for c in range(1, n_cols + 1):
        ws.cell(row=row, column=c).fill = _Z_DRIVER_FILL


def is_z_driver(value: Any) -> bool:
    """True if the cell value looks like the Z driver placeholder."""
    if value is None:
        return False
    return str(value).strip().upper() == "Z"
