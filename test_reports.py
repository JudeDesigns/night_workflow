import sys
import os
import traceback
from pathlib import Path
from openpyxl import load_workbook

# Add project root to path
sys.path.append(os.getcwd())
sys.path.append(str(Path(os.getcwd()) / "backend"))

from nightshift.part1_get_results import compile_all_orders, extract_jetro_and_po
from nightshift.part1_shortages import calculate_shortages
from nightshift.part1_setup import setup_sheets_and_dropdowns
from nightshift.part2_apply_routing import apply_routing  # noqa: F401  (optional)
from nightshift.part3_final_pick_sheets import build_final_pick_sheets
from nightshift.part4_jetro_workflow import build_jetro_branch


def _nonempty(ws, col_idx, header_label):
    """Count data rows where the given column is non-empty (skips title/header rows)."""
    total = 0
    filled = 0
    for row in ws.iter_rows(values_only=True):
        if not any(c not in (None, "") for c in row):
            continue
        first = str(row[0] or "")
        # crude skip of merged title / header rows
        if header_label in [str(c) for c in row]:
            continue
        total += 1
        if col_idx <= len(row) and row[col_idx - 1] not in (None, ""):
            filled += 1
    return filled, total


def run_full_pipeline(job_id):
    job_dir = Path("backend/_jobs") / job_id
    upload = job_dir / "upload.xlsx"
    print(f"--- Full pipeline from upload: {job_id} ---")
    if not upload.exists():
        print("upload.xlsx not found")
        return

    wb = load_workbook(str(upload))
    compile_all_orders(wb)
    extract_jetro_and_po(wb)
    calculate_shortages(wb)
    setup_sheets_and_dropdowns(wb)
    wb.save(str(job_dir / "part1.xlsx"))
    print("Part 1 complete -> part1.xlsx")

    # Use part1 output directly as routing baseline for the test
    wb.save(str(job_dir / "part2.xlsx"))

    build_final_pick_sheets(wb, str(job_dir))
    build_jetro_branch(wb, str(job_dir))
    print("Reports generated.\n")

    # --- Verify the user-reported issues ---
    print("=== VERIFICATION ===")
    pick = load_workbook(str(job_dir / "pick_sheets.xlsx"))
    print("pick_sheets.xlsx sheets:", pick.sheetnames)
    if "Dry by Driver" in pick.sheetnames:
        dry = pick["Dry by Driver"]
        print("  Dry by Driver dims:", dry.max_row, "x", dry.max_column)
        f, t = _nonempty(dry, 1, "Internal bin")
        print(f"  Dry Internal bin filled: {f}/{t}")

    po = load_workbook(str(job_dir / "po_report.xlsx"))
    pov = po["PO by Vendors"]
    f, t = _nonempty(pov, 6, "CURRENT COST PRICE")
    print(f"  PO CURRENT COST PRICE filled: {f}/{t}")

    jetro = load_workbook(str(job_dir / "jetro.xlsx"))
    jp = jetro["Jetro Page"]
    f, t = _nonempty(jp, 1, "Internal bin")
    print(f"  Jetro Page Internal bin filled: {f}/{t}")
    print("  Jetro Page header sample:", [c.value for c in jp[2]])


if __name__ == "__main__":
    try:
        run_full_pipeline("e067e46d")
    except Exception:
        print("\n!!! ERROR DETECTED !!!")
        traceback.print_exc()
