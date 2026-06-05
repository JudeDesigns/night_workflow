import openpyxl
import sys
import os

# Add the project root to sys.path to import local modules
sys.path.append(os.getcwd())

from backend.nightshift.sheet_utils import find_header_row, find_sheet

def diagnose_file(file_path):
    print(f"--- Diagnosing File: {file_path} ---")
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"Error loading workbook: {e}")
        return

    try:
        # Based on constants.py, SHEET_ALL_ORDERS = "All orders Sheet."
        sheet_name = "All orders Sheet." 
        ws = find_sheet(wb, sheet_name)
        print(f"Found sheet: {ws.title}")
    except KeyError as e:
        print(f"Error finding sheet: {e}")
        print(f"Available sheets: {wb.sheetnames}")
        return

    print("\n--- Detailed Inspection of Row 1 ---")
    row1 = [str(cell.value) for cell in ws[1]]
    print(f"Row 1: {row1}")
    
    print("\n--- Detailed Inspection of Rows 2-10 ---")
    for i in range(2, 11):
        row = [str(cell.value) for cell in ws[i]]
        print(f"Row {i}: {row}")

    print("\n--- Testing find_header_row with specific column check ---")
    detected_idx = find_header_row(ws, "Bin", "Quantity On Hand", "Code")
    print(f"Detected Header Row Index (Bin/QOH/Code): {detected_idx}")
    
    # Try another common set of headers
    detected_idx_2 = find_header_row(ws, "Product Name", "Description", "Vendor")
    print(f"Detected Header Row Index (Name/Desc/Vendor): {detected_idx_2}")

    print("\n--- Testing iter_data_rows ---")
    from backend.nightshift.sheet_utils import iter_data_rows
    data_rows = list(iter_data_rows(ws, ["Bin", "Quantity On Hand", "Code"]))
    print(f"Total Data Rows extracted: {len(data_rows)}")
    if data_rows:
        print(f"First Data Row: {data_rows[0][0]}")
        print(f"Last Data Row: {data_rows[-1][0]}")

if __name__ == "__main__":
    file_path = "/Users/mac/code_projects/night_workflow/All orders Sheet (1).xlsx"
    diagnose_file(file_path)
